"""Geração de templates Excel para importação nas bases Notion RPADV.

Produz arquivos .xlsx com cabeçalhos corretos, validação de lista para campos
select e formatação visual básica para facilitar o preenchimento.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from notion_bulk_edit.schemas import SCHEMAS, PropSpec


# ---------------------------------------------------------------------------
# Constantes de estilo
# ---------------------------------------------------------------------------

_COR_HEADER_BG = "1F4E79"      # Azul escuro RPADV
_COR_HEADER_FG = "FFFFFF"      # Branco
_COR_OBRIG_BG  = "FFF2CC"      # Amarelo claro para campos obrigatórios
_COR_READONLY_BG = "F2F2F2"    # Cinza claro para somente leitura
_COR_EDITAVEL_BG = "DEEAF1"    # Azul claro para campos editáveis


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------


def gerar_template(base: str, output_path: str) -> None:
    """Gera um template .xlsx para importação em uma base Notion.

    Colunas incluídas: todas as propriedades editáveis do schema da base.
    Tipos somente leitura (rollup, formula, created_time, last_edited_time)
    são incluídos como referência mas marcados em cinza e protegidos.

    Para campos select e multi_select, adiciona validação de lista (dropdown)
    com as opções do vocabulário controlado do schema.

    Args:
        base: Nome da base ('Processos', 'Clientes', 'Tarefas', 'Catalogo').
        output_path: Caminho do arquivo .xlsx a gerar.

    Raises:
        ValueError: Se a base não existir no schema.
        ImportError: Se openpyxl não estiver instalado.

    Exemplo:
        >>> gerar_template("Processos", "/tmp/template_processos.xlsx")
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise ImportError(
            "Pacote 'openpyxl' não instalado. Execute: pip install openpyxl"
        ) from exc

    schema = SCHEMAS.get(base)
    if schema is None:
        raise ValueError(
            f"Base '{base}' não encontrada. Bases disponíveis: {list(SCHEMAS.keys())}"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = base

    # ------------------------------------------------------------------
    # Aba de referência: vocabulários para os dropdowns
    # ------------------------------------------------------------------

    ws_vocab = wb.create_sheet("_vocabularios")
    ws_vocab.sheet_state = "hidden"

    # Mapa: chave_campo → letra da coluna no _vocabularios
    vocab_col_map: dict[str, str] = {}
    vocab_col_idx = 1

    for key, spec in schema.items():
        if spec.opcoes and spec.tipo in ("select", "multi_select"):
            col_letter = get_column_letter(vocab_col_idx)
            vocab_col_map[key] = col_letter
            # Cabeçalho
            ws_vocab.cell(row=1, column=vocab_col_idx, value=spec.notion_name)
            # Opções
            for row_i, opcao in enumerate(spec.opcoes, start=2):
                ws_vocab.cell(row=row_i, column=vocab_col_idx, value=opcao)
            vocab_col_idx += 1

    # ------------------------------------------------------------------
    # Estilos reutilizáveis
    # ------------------------------------------------------------------

    header_fill     = PatternFill("solid", fgColor=_COR_HEADER_BG)
    header_font     = Font(color=_COR_HEADER_FG, bold=True, size=10)
    obrig_fill      = PatternFill("solid", fgColor=_COR_OBRIG_BG)
    readonly_fill   = PatternFill("solid", fgColor=_COR_READONLY_BG)
    editavel_fill   = PatternFill("solid", fgColor=_COR_EDITAVEL_BG)
    thin_side       = Side(style="thin", color="BFBFBF")
    thin_border     = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align      = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _READONLY_TIPOS = ("rollup", "formula", "created_time", "last_edited_time")

    # ------------------------------------------------------------------
    # Linha 1: cabeçalhos
    # ------------------------------------------------------------------

    # Coluna auxiliar: page_id (necessário para aplicar edições)
    ws.cell(row=1, column=1, value="page_id").font = Font(color="808080", bold=True, size=9)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="E7E6E6")
    ws.cell(row=1, column=1).alignment = center_align
    ws.column_dimensions["A"].width = 36

    # Linhas de dica (linha 2): informações sobre o campo
    ws.cell(row=2, column=1, value="(ID Notion — não editar)").font = Font(color="808080", italic=True, size=8)
    ws.cell(row=2, column=1).alignment = left_align

    # Colunas do schema
    for col_idx, (key, spec) in enumerate(schema.items(), start=2):
        col_letter = get_column_letter(col_idx)

        is_readonly = (not spec.editavel) or (spec.tipo in _READONLY_TIPOS)

        # --- Cabeçalho ---
        header_cell = ws.cell(row=1, column=col_idx, value=spec.notion_name)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center_align
        header_cell.border = thin_border

        # Asterisco para obrigatórios
        if spec.obrigatorio:
            header_cell.value = f"{spec.notion_name} *"

        # --- Linha de dica ---
        dica = _build_dica(spec)
        dica_cell = ws.cell(row=2, column=col_idx, value=dica)
        dica_cell.font = Font(color="595959", italic=True, size=8)
        dica_cell.alignment = left_align
        if is_readonly:
            dica_cell.fill = readonly_fill
        elif spec.obrigatorio:
            dica_cell.fill = obrig_fill
        else:
            dica_cell.fill = editavel_fill

        # --- Largura da coluna ---
        ws.column_dimensions[col_letter].width = _calcular_largura(spec)

        # --- Preenchimento das linhas de dados (3..102) ---
        for data_row in range(3, 103):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = thin_border
            if is_readonly:
                cell.fill = readonly_fill
            cell.alignment = left_align

        # --- Validação de lista para selects ---
        if spec.tipo in ("select", "multi_select") and key in vocab_col_map:
            vocab_col = vocab_col_map[key]
            n_opcoes = len(spec.opcoes)
            formula = f"_vocabularios!${vocab_col}$2:${vocab_col}${n_opcoes + 1}"
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Valor inválido",
                error=f"Selecione uma opção válida para '{spec.label}'.",
            )
            dv.sqref = f"{col_letter}3:{col_letter}102"
            ws.add_data_validation(dv)

        # --- Validação de data ---
        if spec.tipo == "date":
            dv_date = DataValidation(
                type="custom",
                formula1=(
                    f'OR(AND(LEN({col_letter}3)=10,MID({col_letter}3,5,1)="-",MID({col_letter}3,8,1)="-"),'
                    f'AND(LEN({col_letter}3)=10,MID({col_letter}3,3,1)="/",MID({col_letter}3,6,1)="/"))'
                ),
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Formato de data inválido",
                error="Use YYYY-MM-DD ou DD/MM/YYYY.",
            )
            dv_date.sqref = f"{col_letter}3:{col_letter}102"
            ws.add_data_validation(dv_date)

    # ------------------------------------------------------------------
    # Linha de instrução (linha 1 — sobrepõe cabeçalho da aba com comentário)
    # ------------------------------------------------------------------

    # Congela a primeira linha de cabeçalho + linha de dica
    ws.freeze_panes = "B3"

    # Altura das linhas especiais
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # ------------------------------------------------------------------
    # Aba de instruções
    # ------------------------------------------------------------------

    _adicionar_aba_instrucoes(wb, base, schema)

    # ------------------------------------------------------------------
    # Salvar
    # ------------------------------------------------------------------

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"Template gerado: {output}")


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------


def _build_dica(spec: PropSpec) -> str:
    """Constrói a string de dica para a linha 2 do template.

    Args:
        spec: PropSpec da propriedade.

    Returns:
        String descritiva do campo.
    """
    partes: list[str] = [spec.tipo]

    if spec.obrigatorio:
        partes.append("obrigatório")

    if not spec.editavel:
        partes.append("somente leitura")

    if spec.opcoes:
        opcoes_str = ", ".join(spec.opcoes[:5])
        if len(spec.opcoes) > 5:
            opcoes_str += f"... (+{len(spec.opcoes) - 5})"
        partes.append(f"Opções: {opcoes_str}")

    if spec.formato == "BR_DATE":
        partes.append("Ex: 31/12/2024 ou 2024-12-31")
    elif spec.formato == "BRL":
        partes.append("Ex: 78500.00")

    return " | ".join(partes)


def _calcular_largura(spec: PropSpec) -> float:
    """Calcula uma largura razoável para a coluna em caracteres.

    Args:
        spec: PropSpec da propriedade.

    Returns:
        Largura em unidades do openpyxl.
    """
    largura_map = {
        "title":        40,
        "rich_text":    30,
        "number":       14,
        "select":       18,
        "multi_select": 22,
        "date":         14,
        "people":       16,
        "checkbox":     10,
        "relation":     36,
        "rollup":       16,
        "formula":      16,
        "url":          30,
        "email":        26,
        "phone_number": 16,
        "created_time": 22,
        "last_edited_time": 22,
    }
    return largura_map.get(spec.tipo, 18)


def _adicionar_aba_instrucoes(wb: Any, base: str, schema: dict[str, PropSpec]) -> None:
    """Adiciona uma aba de instruções ao workbook.

    Args:
        wb: Workbook openpyxl.
        base: Nome da base.
        schema: Schema da base.
    """
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet("Instruções")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    titulo_font = Font(bold=True, size=13, color="1F4E79")
    subtitulo_font = Font(bold=True, size=10)
    normal_font = Font(size=10)

    linhas: list[tuple[str, str, Any]] = [
        ("NOTION RPADV", f"Template de importação — Base: {base}", titulo_font),
        ("", "", normal_font),
        ("Instruções:", "", subtitulo_font),
        ("1.", "Preencha os dados a partir da linha 3.", normal_font),
        ("2.", "Campos com * no cabeçalho são obrigatórios.", normal_font),
        ("3.", "A coluna 'page_id' deve conter o ID da página Notion para atualizar registros existentes.", normal_font),
        ("4.", "Para criar novos registros, deixe 'page_id' em branco.", normal_font),
        ("5.", "Datas: use YYYY-MM-DD ou DD/MM/YYYY.", normal_font),
        ("6.", "Campos em cinza são somente leitura — não serão importados.", normal_font),
        ("", "", normal_font),
        ("Campos:", "", subtitulo_font),
    ]

    for key, spec in schema.items():
        obrig = " (obrigatório)" if spec.obrigatorio else ""
        readonly = " [somente leitura]" if not spec.editavel else ""
        descricao = f"{spec.label}{obrig}{readonly} — tipo: {spec.tipo}"
        if spec.opcoes:
            descricao += f" — opções: {', '.join(spec.opcoes)}"
        linhas.append((spec.notion_name, descricao, normal_font))

    for row_i, (col_a, col_b, font) in enumerate(linhas, start=1):
        cell_a = ws.cell(row=row_i, column=1, value=col_a)
        cell_b = ws.cell(row=row_i, column=2, value=col_b)
        cell_a.font = font
        cell_b.font = font
        cell_a.alignment = Alignment(vertical="center")
        cell_b.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_i].height = 16


# ---------------------------------------------------------------------------
# CLI mínimo
# ---------------------------------------------------------------------------


def _main() -> None:
    """Ponto de entrada: python -m notion_bulk_edit.gerar_template BASE SAIDA.xlsx"""
    import sys

    if len(sys.argv) < 3:
        print("Uso: python -m notion_bulk_edit.gerar_template BASE SAIDA.xlsx")
        print(f"Bases disponíveis: {list(SCHEMAS.keys())}")
        sys.exit(1)

    base = sys.argv[1]
    saida = sys.argv[2]

    try:
        gerar_template(base, saida)
    except (ValueError, ImportError) as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    _main()
