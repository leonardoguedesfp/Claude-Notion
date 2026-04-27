#!/usr/bin/env python3
"""CLI legado — edição em massa das bases Notion via linha de comando.

Uso:
  python -m notion_bulk_edit --base Processos --planilha arquivo.xlsx --dry-run
  python -m notion_bulk_edit --base Processos --planilha arquivo.xlsx --aplicar
  python -m notion_bulk_edit --base Processos --exportar saida.xlsx
  python -m notion_bulk_edit --listar-bases
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Importações internas (lazy quando possível para startup rápido)
# ---------------------------------------------------------------------------

from notion_bulk_edit.config import DATA_SOURCES
from notion_bulk_edit.schemas import SCHEMAS, PropSpec
from notion_bulk_edit.encoders import decode_value, encode_value, format_br_date, format_brl
from notion_bulk_edit.validators import validar_linha, ValidationError
from notion_rpadv.auth.token_store import get_token


# ---------------------------------------------------------------------------
# Helpers de saída
# ---------------------------------------------------------------------------


def _print_erros(erros: list[ValidationError], row_num: int) -> None:
    """Imprime os erros de validação de uma linha no stderr."""
    for e in erros:
        print(f"  Linha {row_num}: [{e.campo}] {e.mensagem}", file=sys.stderr)


def _print_ok(msg: str) -> None:
    print(f"  OK  {msg}")


def _print_skip(msg: str) -> None:
    print(f"  --  {msg} (sem alterações)")


def _print_erro(msg: str) -> None:
    print(f"  ERR {msg}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Leitura de planilha XLSX
# ---------------------------------------------------------------------------


def _ler_planilha(path: str, base: str) -> list[dict[str, Any]]:
    """Lê uma planilha XLSX e retorna lista de dicts com os dados.

    A primeira linha deve conter os cabeçalhos. Os cabeçalhos devem
    corresponder aos notion_name das propriedades do schema da base.

    Args:
        path: Caminho para o arquivo .xlsx.
        base: Nome da base para mapear os cabeçalhos.

    Returns:
        Lista de dicts: [{notion_name: valor, ...}, ...]

    Raises:
        FileNotFoundError: Se o arquivo não existir.
        ValueError: Se a planilha não tiver cabeçalhos reconhecíveis.
    """
    try:
        import openpyxl
    except ImportError:
        print("Erro: pacote 'openpyxl' não instalado. Execute: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    caminho = Path(path)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        cabecalhos = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        raise ValueError("Planilha vazia.")

    linhas: list[dict[str, Any]] = []
    for row in rows_iter:
        # Pula linhas completamente vazias
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d: dict[str, Any] = {}
        for header, cell in zip(cabecalhos, row):
            if header:
                d[header] = cell
        linhas.append(d)

    wb.close()
    return linhas


# ---------------------------------------------------------------------------
# Exportação para XLSX
# ---------------------------------------------------------------------------


def _exportar_xlsx(base: str, client: Any, output_path: str) -> None:
    """Exporta todos os registros de uma base para uma planilha XLSX.

    Args:
        base: Nome da base ('Processos', 'Clientes', 'Tarefas', 'Catalogo').
        client: Instância de NotionClient autenticada.
        output_path: Caminho do arquivo de saída.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Erro: pacote 'openpyxl' não instalado. Execute: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    db_id = DATA_SOURCES.get(base)
    if not db_id:
        print(f"Erro: base '{base}' não configurada em DATA_SOURCES.", file=sys.stderr)
        sys.exit(1)

    schema = SCHEMAS.get(base, {})
    colunas = list(schema.keys())

    print(f"Consultando base '{base}'...")

    total_lido = 0

    def progresso(n: int) -> None:
        nonlocal total_lido
        total_lido = n
        print(f"  {n} registros lidos...", end="\r")

    pages = client.query_all(db_id, on_progress=progresso)
    print(f"\n  Total: {len(pages)} registros.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = base

    # Cabeçalho
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, chave in enumerate(colunas, start=1):
        spec: PropSpec = schema[chave]
        cell = ws.cell(row=1, column=col_idx, value=spec.notion_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, page in enumerate(pages, start=2):
        props = page.get("properties", {})
        for col_idx, chave in enumerate(colunas, start=1):
            spec = schema[chave]
            prop_block = props.get(spec.notion_name, {})
            valor = decode_value(prop_block, spec.tipo)

            # Formatar valores para exibição legível
            if spec.formato == "BR_DATE" and isinstance(valor, str):
                valor = format_br_date(valor)
            elif spec.formato == "BRL" and isinstance(valor, (int, float)):
                valor = format_brl(valor)
            elif isinstance(valor, list):
                valor = ", ".join(str(v) for v in valor)

            ws.cell(row=row_idx, column=col_idx, value=valor)

    wb.save(output_path)
    print(f"Exportação concluída: {output_path}")


# ---------------------------------------------------------------------------
# Importação / aplicação de edições
# ---------------------------------------------------------------------------


def _importar(
    base: str,
    planilha: str,
    client: Any,
    dry_run: bool,
) -> None:
    """Valida e aplica as alterações de uma planilha numa base Notion.

    Em modo --dry-run nenhuma requisição PATCH é enviada; apenas valida
    e mostra o que seria feito.

    Args:
        base: Nome da base alvo.
        planilha: Caminho do arquivo .xlsx.
        client: Instância de NotionClient autenticada.
        dry_run: Se True, apenas valida sem aplicar.
    """
    schema = SCHEMAS.get(base, {})
    db_id = DATA_SOURCES.get(base)

    if not db_id:
        print(f"Erro: base '{base}' não configurada em DATA_SOURCES.", file=sys.stderr)
        sys.exit(1)

    print(f"Lendo planilha: {planilha}")
    linhas = _ler_planilha(planilha, base)
    print(f"  {len(linhas)} linha(s) encontrada(s).")

    # Monta mapa notion_name → chave do schema
    notion_name_to_key: dict[str, str] = {
        spec.notion_name: key for key, spec in schema.items()
    }

    erros_total = 0
    aplicados = 0
    ignorados = 0

    for idx, row in enumerate(linhas, start=2):
        erros = validar_linha(base, row)
        if erros:
            erros_total += len(erros)
            _print_erros(erros, idx)
            continue

        # Monta payload de propriedades para a API
        properties: dict[str, Any] = {}
        for col_name, valor in row.items():
            # Resolve chave do schema
            chave = col_name if col_name in schema else notion_name_to_key.get(col_name)
            if chave is None:
                continue
            spec: PropSpec = schema[chave]

            # Só edita campos editáveis e não somente leitura
            if not spec.editavel or spec.tipo in ("rollup", "formula", "created_time", "last_edited_time"):
                continue

            # Não encoda valores vazios (preserva dado existente)
            if valor is None or str(valor).strip() == "":
                continue

            payload = encode_value(valor, spec.tipo)
            if payload:
                properties[spec.notion_name] = payload

        if not properties:
            ignorados += 1
            _print_skip(f"Linha {idx} sem alterações detectadas.")
            continue

        # Determina o page_id — espera coluna "page_id" na planilha
        page_id = row.get("page_id") or row.get("Page ID") or row.get("ID")

        if dry_run:
            print(f"  DRY Linha {idx}: atualizaria {len(properties)} campo(s).")
            if page_id:
                print(f"      page_id: {page_id}")
            aplicados += 1
        else:
            if not page_id:
                _print_erro(
                    f"Linha {idx}: coluna 'page_id' ausente — não é possível aplicar. "
                    "Exporte a base primeiro para obter os IDs."
                )
                erros_total += 1
                continue
            try:
                client.update_page(str(page_id), properties)
                _print_ok(f"Linha {idx}: {len(properties)} campo(s) atualizado(s).")
                aplicados += 1
            except Exception as exc:
                _print_erro(f"Linha {idx}: falha ao atualizar — {exc}")
                erros_total += 1

    # Resumo
    print()
    print("=" * 50)
    print(f"Base:     {base}")
    print(f"Linhas:   {len(linhas)}")
    print(f"Aplicadas:{aplicados}")
    print(f"Ignoradas:{ignorados}")
    print(f"Erros:    {erros_total}")
    if dry_run:
        print("[DRY RUN] Nenhuma alteração foi salva no Notion.")
    print("=" * 50)

    if erros_total:
        sys.exit(1)


# ---------------------------------------------------------------------------
# Subcomandos auxiliares
# ---------------------------------------------------------------------------


def _listar_bases() -> None:
    """Imprime as bases disponíveis e seus IDs configurados."""
    print("Bases disponíveis:")
    for nome, db_id in DATA_SOURCES.items():
        schema = SCHEMAS.get(nome, {})
        n_campos = len(schema)
        print(f"  {nome:<12} | {n_campos} campo(s) | ID: {db_id}")


# ---------------------------------------------------------------------------
# Parser de argumentos
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m notion_bulk_edit",
        description="CLI legado — edição em massa das bases Notion RPADV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    parser.add_argument(
        "--base",
        choices=list(DATA_SOURCES.keys()),
        help="Base Notion a operar (Processos, Clientes, Tarefas, Catalogo).",
    )
    parser.add_argument(
        "--planilha",
        metavar="ARQUIVO.xlsx",
        help="Planilha Excel de importação/edição.",
    )
    parser.add_argument(
        "--exportar",
        metavar="SAIDA.xlsx",
        help="Exporta todos os registros da base para o arquivo indicado.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida a planilha sem aplicar alterações no Notion.",
    )
    parser.add_argument(
        "--aplicar",
        action="store_true",
        help="Aplica as alterações da planilha no Notion.",
    )
    parser.add_argument(
        "--listar-bases",
        action="store_true",
        help="Lista as bases disponíveis e seus IDs configurados.",
    )

    return parser


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> None:
    """Ponto de entrada do CLI legado.

    Args:
        argv: Lista de argumentos (None usa sys.argv).
    """
    parser = _build_parser()
    args = parser.parse_args(argv)

    # --listar-bases não precisa de token
    if args.listar_bases:
        _listar_bases()
        return

    # Qualquer outro comando requer --base
    if not args.base:
        parser.error("--base é obrigatório para este comando.")

    # Verificar ações mutuamente exclusivas
    acao_importar = args.dry_run or args.aplicar
    acao_exportar = bool(args.exportar)

    if acao_importar and acao_exportar:
        parser.error("Não é possível usar --exportar e --planilha/--dry-run/--aplicar simultaneamente.")

    if acao_importar and not args.planilha:
        parser.error("--planilha é obrigatório para --dry-run ou --aplicar.")

    if not acao_importar and not acao_exportar:
        parser.error("Informe --dry-run, --aplicar ou --exportar.")

    # Obter token
    token = get_token()
    if not token:
        print(
            "Erro: token Notion não configurado.\n"
            "Configure via o app desktop ou:\n"
            "  python -c \"from notion_rpadv.auth.token_store import set_token; set_token('secret_...')\"",
            file=sys.stderr,
        )
        sys.exit(1)

    # Instanciar cliente (import lazy)
    from notion_bulk_edit.notion_api import NotionClient, NotionAuthError

    client = NotionClient(token)

    # Validar token
    try:
        client.me()
    except NotionAuthError as exc:
        print(f"Erro de autenticação: {exc}", file=sys.stderr)
        sys.exit(1)

    # Executar ação
    if acao_exportar:
        _exportar_xlsx(args.base, client, args.exportar)
    else:
        _importar(
            base=args.base,
            planilha=args.planilha,
            client=client,
            dry_run=args.dry_run,
        )


if __name__ == "__main__":
    main()
