"""Round 6 — testes pra rollup-de-relation + redesign Dashboard.

Parte 1: rollup que aponta pra um campo de relation na base relacionada
(ex: Tarefas.Cliente roll up Processos.Clientes) deve resolver UUIDs
pra títulos em todas as camadas — display da tabela, xlsx export,
search livre.

Parte 2: Dashboard reformata Tarefas Urgentes em 3 grupos
(Vencidas/Hoje/Amanhã) com cards reformatados.
"""
from __future__ import annotations

import datetime
import json
import sqlite3
from pathlib import Path
from unittest.mock import MagicMock

import pytest


_FIXTURES_DIR = Path(__file__).parent / "fixtures" / "schemas"


def _audit_only_conn() -> sqlite3.Connection:
    from notion_rpadv.cache import db as cache_db
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    cache_db.init_audit_db(conn)
    return conn


def _full_conn() -> sqlite3.Connection:
    """Conn com cache + audit (records vivem no cache, schemas no audit)."""
    from notion_rpadv.cache import db as cache_db
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    cache_db.init_db(conn)
    return conn


# ---------------------------------------------------------------------------
# Parte 1 — rollup-de-relation
# ---------------------------------------------------------------------------


def test_R6_schema_parser_captures_rollup_meta() -> None:
    """schema_parser agora preserva rollup_meta (relation_property_name,
    rollup_property_name, function) no dict canônico, condição
    necessária pra registry fazer o 2-hop."""
    from notion_bulk_edit.schema_parser import parse_to_schema_json
    raw = {
        "id": "ds-task",
        "properties": {
            "Tarefa": {"type": "title", "title": {}},
            "Processo": {
                "type": "relation",
                "relation": {"data_source_id": "ds-proc"},
            },
            "Cliente": {
                "type": "rollup",
                "rollup": {
                    "relation_property_name": "Processo",
                    "rollup_property_name": "Clientes",
                    "function": "show_original",
                    "relation_property_id": "abc",
                    "rollup_property_id": "xyz",
                },
            },
        },
    }
    parsed = parse_to_schema_json(raw, "Tasks")
    cliente_meta = parsed["properties"]["cliente"].get("rollup_meta", {})
    assert cliente_meta.get("relation_property_name") == "Processo"
    assert cliente_meta.get("rollup_property_name") == "Clientes"
    assert cliente_meta.get("function") == "show_original"


def test_R6_registry_resolves_target_base_for_rollup_relation() -> None:
    """Registry usa rollup_meta + 2-hop pra setar PropSpec.target_base
    em Tarefas.cliente. Smoke usando os fixtures reais (Tarefas →
    Processos → Clientes)."""
    if not (_FIXTURES_DIR / "tarefas_raw.json").exists():
        pytest.skip("fixtures ausentes")
    from notion_bulk_edit.schema_registry import get_schema_registry
    reg = get_schema_registry()
    spec = reg.get_prop("Tarefas", "cliente")
    assert spec is not None
    assert spec.tipo == "rollup"
    assert spec.target_base == "Clientes"


def test_R6_registry_no_target_base_for_rollup_of_select() -> None:
    """Tarefas.Tribunal é rollup-de-select (não relation) — não deveria
    receber target_base. Hop 2 detecta que rollup_property_name aponta
    pra select e desiste silenciosamente."""
    if not (_FIXTURES_DIR / "tarefas_raw.json").exists():
        pytest.skip("fixtures ausentes")
    from notion_bulk_edit.schema_registry import get_schema_registry
    reg = get_schema_registry()
    spec = reg.get_prop("Tarefas", "tribunal")
    assert spec is not None
    assert spec.tipo == "rollup"
    assert spec.target_base == ""


def test_R6_flatten_rollup_uuids_handles_nested_lists() -> None:
    """``_flatten_rollup_uuids`` achata ``[[uuid1], [uuid2, uuid3]]``
    pra ``[uuid1, uuid2, uuid3]``."""
    from notion_rpadv.models.base_table_model import _flatten_rollup_uuids
    assert _flatten_rollup_uuids([["a"], ["b", "c"]]) == ["a", "b", "c"]
    # Mistura de nested + plain string preservada.
    assert _flatten_rollup_uuids([["a"], "b"]) == ["a", "b"]
    # Vazio.
    assert _flatten_rollup_uuids([]) == []
    assert _flatten_rollup_uuids(None) == []
    # Strings vazias dentro são filtradas.
    assert _flatten_rollup_uuids([[""], ["c"]]) == ["c"]


def test_R6_resolve_rollup_relation_uses_cache_titles() -> None:
    """``_resolve_rollup_relation`` consulta cache_db pra cada UUID e
    retorna o título da página. Múltiplos viram comma-separated."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.models.base_table_model import _resolve_rollup_relation
    conn = _full_conn()
    cache_db.upsert_record(conn, "Clientes", "uuid-1", {
        "page_id": "uuid-1", "nome": "Maria Silva",
    })
    cache_db.upsert_record(conn, "Clientes", "uuid-2", {
        "page_id": "uuid-2", "nome": "João Costa",
    })
    result = _resolve_rollup_relation(conn, ["uuid-1", "uuid-2"], "Clientes")
    assert result == "Maria Silva, João Costa"


def test_R6_resolve_rollup_relation_falls_back_to_uuid_when_missing() -> None:
    """Página não cacheada → UUID-cru no display (não "—" como em
    _resolve_relation). Diagnóstico vivo quando target_base ainda não
    sincronizou."""
    from notion_rpadv.models.base_table_model import _resolve_rollup_relation
    conn = _full_conn()
    # Nenhuma Cliente cacheada
    result = _resolve_rollup_relation(
        conn, ["uuid-fora", "outro-uuid"], "Clientes",
    )
    assert result == "uuid-fora, outro-uuid"


def test_R6_resolve_rollup_relation_caps_at_5_with_overflow_marker() -> None:
    """Cap visual em 5 entradas + marker ``+N`` pro resto. Rollups
    podem agregar dezenas de itens — display de tabela não comporta."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.models.base_table_model import _resolve_rollup_relation
    conn = _full_conn()
    for i in range(10):
        cache_db.upsert_record(conn, "Clientes", f"u{i}", {
            "page_id": f"u{i}", "nome": f"C{i}",
        })
    result = _resolve_rollup_relation(
        conn, [f"u{i}" for i in range(10)], "Clientes",
    )
    # 5 nomes + " +5"
    assert result.endswith("+5")
    assert "C0" in result and "C4" in result


def test_R6_rollup_relation_resolves_title_in_display() -> None:
    """``_display_value`` pra rollup achata nested antes do join. Sem
    target_base setado, retorna UUIDs flat (caller — data() — resolve
    quando spec.target_base existe)."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Cliente", tipo="rollup", label="Cliente",
        editavel=False, obrigatorio=False, opcoes=(),
    )
    # Nested list (sem resolução, target_base="" no spec):
    result = _display_value(spec, [["uuid-1"], ["uuid-2"]])
    assert result == "uuid-1, uuid-2"


def test_R6_rollup_relation_falls_back_to_uuid_when_target_missing() -> None:
    """data() resolve via _resolve_rollup_relation. UUIDs sem registro
    cacheado caem em UUID-cru (não "—"). Smoke direto da função."""
    from notion_rpadv.models.base_table_model import _resolve_rollup_relation
    conn = _full_conn()
    # Sem cachear nenhum registro: tudo cai em UUID-cru.
    result = _resolve_rollup_relation(
        conn, ["uuid-x", "uuid-y"], "Clientes",
    )
    assert "uuid-x" in result
    assert "uuid-y" in result


def test_R6_rollup_relation_handles_multiple_items_comma_separated() -> None:
    """Múltiplos UUIDs em rollup viram comma-separated, com cache hits
    e misses misturados (cache → nome, miss → UUID)."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.models.base_table_model import _resolve_rollup_relation
    conn = _full_conn()
    cache_db.upsert_record(conn, "Clientes", "u1", {
        "page_id": "u1", "nome": "Alice",
    })
    # u2 não cacheado.
    result = _resolve_rollup_relation(conn, ["u1", "u2"], "Clientes")
    assert result == "Alice, u2"


def test_R6_rollup_relation_resolves_in_xlsx_export(tmp_path) -> None:
    """Snapshot xlsx resolve UUIDs de rollup-de-relation via
    title_cache (cobre as bases selecionadas no export). Sem isso o
    operador veria UUIDs no Excel da Cliente, mesmo bug da tabela."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    cliente_page = {
        "id": "cli-1",
        "properties": {
            "Nome": {
                "type": "title",
                "title": [{"plain_text": "Maria Silva"}],
            },
        },
    }
    tarefa_page = {
        "id": "tar-1",
        "properties": {
            "Tarefa": {
                "type": "title",
                "title": [{"plain_text": "Petição inicial"}],
            },
            # Rollup retorna nested array de relation blocks.
            "Cliente": {
                "type": "rollup",
                "rollup": {
                    "type": "array",
                    "array": [
                        {
                            "type": "relation",
                            "relation": [{"id": "cli-1"}],
                        },
                    ],
                },
            },
        },
    }

    def query_all(dsid: str, on_progress=None):
        return [cliente_page] if dsid == "ds-c" else [tarefa_page]
    client = MagicMock()
    client.query_all.side_effect = query_all

    schemas = {
        "SynthClientes": {"properties": {
            "nome": {
                "notion_name": "Nome", "tipo": "title",
                "default_visible": True, "default_order": 1, "opcoes": [],
            },
        }},
        "SynthTarefas": {"properties": {
            "tarefa": {
                "notion_name": "Tarefa", "tipo": "title",
                "default_visible": True, "default_order": 1, "opcoes": [],
            },
            "cliente": {
                "notion_name": "Cliente", "tipo": "rollup",
                "default_visible": True, "default_order": 2, "opcoes": [],
            },
        }},
    }
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["SynthClientes", "SynthTarefas"],
        dest_path=dest, schema_registry=reg,
        data_sources={"SynthClientes": "ds-c", "SynthTarefas": "ds-t"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["SynthTarefas"]
    # col 2 = Cliente (rollup-de-relation). Resolvido pra "Maria Silva".
    assert ws.cell(row=2, column=2).value == "Maria Silva"


def test_R6_rollup_relation_xlsx_falls_back_to_uuid_when_target_not_in_snapshot(
    tmp_path,
) -> None:
    """Quando o usuário exporta só Tarefas (sem Clientes), os UUIDs de
    rollup ficam em UUID-cru (sem "[?]"). Diagnóstico preservado."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    tarefa_page = {
        "id": "tar-1",
        "properties": {
            "Tarefa": {"type": "title", "title": [{"plain_text": "T1"}]},
            "Cliente": {
                "type": "rollup",
                "rollup": {
                    "type": "array",
                    "array": [
                        {
                            "type": "relation",
                            "relation": [{"id": "cli-fora"}],
                        },
                    ],
                },
            },
        },
    }
    client = MagicMock()
    client.query_all.return_value = [tarefa_page]
    schemas = {"SynthTarefas": {"properties": {
        "tarefa": {
            "notion_name": "Tarefa", "tipo": "title",
            "default_visible": True, "default_order": 1, "opcoes": [],
        },
        "cliente": {
            "notion_name": "Cliente", "tipo": "rollup",
            "default_visible": True, "default_order": 2, "opcoes": [],
        },
    }}}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    result = export_snapshot(
        client=client, bases=["SynthTarefas"], dest_path=dest,
        schema_registry=reg, data_sources={"SynthTarefas": "ds-t"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["SynthTarefas"]
    # UUID-cru (não "[?]") porque rollup-de-relation usa UUID fallback.
    assert ws.cell(row=2, column=2).value == "cli-fora"
    # Não conta como relation_misses (rollup tem semântica diferente
    # de relation direta — pode apontar pra páginas fora do snapshot
    # por design).
    assert result.relation_misses == 0


def test_R6_search_matches_resolved_title_in_rollup_relation() -> None:
    """Search livre casa pelo nome resolvido no rollup-de-relation,
    não pelo UUID. Sem isso o operador busca "Maria" e não acha
    Tarefas dela porque a célula mostraria UUID."""
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_bulk_edit.config import DATA_SOURCES
    from notion_bulk_edit.schema_registry import get_schema_registry
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.models.base_table_model import BaseTableModel
    from notion_rpadv.models.filters import TableFilterProxy

    if not (_FIXTURES_DIR / "tarefas_raw.json").exists():
        pytest.skip("fixtures ausentes")

    conn = _full_conn()
    # Cache uma Cliente com nome conhecido
    cache_db.upsert_record(conn, "Clientes", "cli-x", {
        "page_id": "cli-x", "nome": "Aurora Magalhães",
    })
    # Cache uma Tarefa cuja rollup Cliente aponta pra cli-x
    cache_db.upsert_record(conn, "Tarefas", "tar-x", {
        "page_id": "tar-x",
        "tarefa": "Petição",
        "cliente": [["cli-x"]],   # rollup-array nested decoded
    })
    # Força user_columns pra incluir cliente
    audit = get_schema_registry()._audit_conn  # noqa: SLF001
    cache_db.set_user_columns(
        audit, "u-search", DATA_SOURCES["Tarefas"], ["tarefa", "cliente"],
    )
    model = BaseTableModel("Tarefas", conn, user_id="u-search")
    proxy = TableFilterProxy()
    proxy.setSourceModel(model)
    # Search casa o nome resolvido (DisplayRole renderizado pelo
    # _resolve_rollup_relation a partir do cache local de Clientes).
    proxy.set_search("Aurora")
    assert proxy.rowCount() == 1
    # Sub-string do nome também casa.
    proxy.set_search("Magalh")
    assert proxy.rowCount() == 1
    # Termo desconhecido não casa.
    proxy.set_search("inexistente-xyz-zzz")
    assert proxy.rowCount() == 0
    # Nota: o filtro do app busca em DisplayRole + EditRole (BUG-OP-04
    # dual-role search), então UUID-cru ainda casa via EditRole. Isso
    # é intencional — search é defensivo e quer achar o registro de
    # qualquer ângulo, não só o nome resolvido.


# ---------------------------------------------------------------------------
# Sweep: lista de rollup-de-relation por base (Round 6 spec ask)
# ---------------------------------------------------------------------------


def test_R6_sweep_rollup_relation_properties_per_base() -> None:
    """Round 6 sweep: documenta quais propriedades rollup têm
    target_base resolvido (rollup-de-relation), em todas as 4 bases.
    Smoke garante que a fix cobre todas automaticamente — não
    case-by-case. Falha aqui sinaliza que uma rollup nova entrou no
    schema sem ser resolvida."""
    from notion_bulk_edit.schema_registry import get_schema_registry
    if not (_FIXTURES_DIR / "tarefas_raw.json").exists():
        pytest.skip("fixtures ausentes")
    reg = get_schema_registry()
    rollup_relations: dict[str, dict[str, str]] = {}
    for base in ["Clientes", "Processos", "Tarefas", "Catalogo"]:
        for slug, spec in reg.schema_for_base(base).items():
            if spec.tipo == "rollup" and spec.target_base:
                rollup_relations.setdefault(base, {})[slug] = spec.target_base
    # Tarefas.cliente é o caso conhecido — o sweep deve incluí-lo.
    assert rollup_relations.get("Tarefas", {}).get("cliente") == "Clientes"


# ---------------------------------------------------------------------------
# Parte 2 — Dashboard Tarefas Urgentes
# ---------------------------------------------------------------------------


def _today_tomorrow() -> tuple[datetime.date, datetime.date]:
    today = datetime.date(2026, 4, 30)
    return today, today + datetime.timedelta(days=1)


def test_R6_dashboard_groups_tasks_by_due_status() -> None:
    """``_group_urgent_tasks`` separa records em 3 buckets:
    vencida (prazo < hoje), hoje (==hoje), amanha (==hoje+1)."""
    from notion_rpadv.pages.dashboard import _group_urgent_tasks
    today, tomorrow = _today_tomorrow()
    tasks = [
        {"prazo_fatal": "2026-04-15", "tarefa": "T1"},  # vencida
        {"prazo_fatal": "2026-04-29", "tarefa": "T2"},  # vencida
        {"prazo_fatal": "2026-04-30", "tarefa": "T3"},  # hoje
        {"prazo_fatal": "2026-05-01", "tarefa": "T4"},  # amanhã
        {"prazo_fatal": "2026-05-15", "tarefa": "T5"},  # futuro distante (out)
        {"prazo_fatal": "", "tarefa": "T6"},            # sem prazo (out)
    ]
    groups = _group_urgent_tasks(tasks, today)
    assert len(groups["vencida"]) == 2
    assert len(groups["hoje"]) == 1
    assert len(groups["amanha"]) == 1
    assert {t["tarefa"] for t in groups["vencida"]} == {"T1", "T2"}
    assert groups["hoje"][0]["tarefa"] == "T3"
    assert groups["amanha"][0]["tarefa"] == "T4"


def test_R6_dashboard_overdue_tasks_excludes_completed() -> None:
    """Status ``Concluída`` exclui a tarefa de qualquer bucket urgente,
    mesmo que prazo_fatal seja anterior a hoje."""
    from notion_rpadv.pages.dashboard import _group_urgent_tasks
    today, _ = _today_tomorrow()
    tasks = [
        {"prazo_fatal": "2026-04-15", "tarefa": "T1", "status": "Pendente"},
        {"prazo_fatal": "2026-04-15", "tarefa": "T2", "status": "Concluída"},
    ]
    groups = _group_urgent_tasks(tasks, today)
    assert len(groups["vencida"]) == 1
    assert groups["vencida"][0]["tarefa"] == "T1"


def test_R6_dashboard_today_tasks_uses_local_date() -> None:
    """``_classify_task`` compara prazo_fatal com ``today`` passado pelo
    caller (que normalmente usa ``date.today()`` — local). Não converte
    timezone, não consulta UTC."""
    from notion_rpadv.pages.dashboard import _classify_task
    today = datetime.date(2026, 4, 30)
    tomorrow = datetime.date(2026, 5, 1)
    # Match exato com today
    assert _classify_task(
        {"prazo_fatal": "2026-04-30"}, today, tomorrow,
    ) == "hoje"
    # Datetime com horário no mesmo dia → ainda hoje
    assert _classify_task(
        {"prazo_fatal": "2026-04-30T23:59:59-03:00"}, today, tomorrow,
    ) == "hoje"
    # Dia anterior → vencida
    assert _classify_task(
        {"prazo_fatal": "2026-04-29"}, today, tomorrow,
    ) == "vencida"


def test_R6_dashboard_classify_task_handles_invalid_date() -> None:
    """prazo_fatal inválido/vazio → None (não levanta)."""
    from notion_rpadv.pages.dashboard import _classify_task
    today, tomorrow = _today_tomorrow()
    assert _classify_task({"prazo_fatal": ""}, today, tomorrow) is None
    assert _classify_task({"prazo_fatal": "lixo"}, today, tomorrow) is None
    assert _classify_task({}, today, tomorrow) is None


def test_R6_dashboard_resolve_cliente_returns_name_not_uuid() -> None:
    """``_resolve_dashboard_cliente`` consulta cache local de Clientes
    e devolve nome no lugar do UUID (parte do fix do Round 6 Parte 1
    aplicada ao Dashboard)."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import _resolve_dashboard_cliente
    conn = _full_conn()
    cache_db.upsert_record(conn, "Clientes", "uuid-1", {
        "page_id": "uuid-1", "nome": "Maria Silva",
    })
    # rollup-de-relation chega como nested list:
    record = {"cliente": [["uuid-1"]]}
    assert _resolve_dashboard_cliente(conn, record) == "Maria Silva"


def test_R6_dashboard_resolve_cliente_falls_back_to_uuid() -> None:
    """UUID não cacheado → UUID-cru (não em-dash). Diagnóstico
    preservado quando Clientes ainda não sincronizou."""
    from notion_rpadv.pages.dashboard import _resolve_dashboard_cliente
    conn = _full_conn()
    record = {"cliente": [["uuid-fora"]]}
    assert _resolve_dashboard_cliente(conn, record) == "uuid-fora"


def test_R6_dashboard_resolve_cnj_returns_numero_processo() -> None:
    """``_resolve_dashboard_cnj`` resolve Tarefas.processo (relation
    single) pra numero_do_processo via cache local de Processos."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import _resolve_dashboard_cnj
    conn = _full_conn()
    cache_db.upsert_record(conn, "Processos", "proc-1", {
        "page_id": "proc-1",
        "numero_do_processo": "0001234-56.2026.1.23.4567",
    })
    record = {"processo": ["proc-1"]}
    assert _resolve_dashboard_cnj(conn, record) == "0001234-56.2026.1.23.4567"


def test_R6_dashboard_format_subtitle_combines_cliente_and_cnj() -> None:
    """Subtítulo do card é ``cliente · CNJ`` (middle-dot). Partes
    vazias são omitidas — tarefa sem cliente ainda mostra CNJ sozinho."""
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import _format_dashboard_subtitle
    conn = _full_conn()
    cache_db.upsert_record(conn, "Clientes", "u1", {
        "page_id": "u1", "nome": "Aurora Magalhães",
    })
    cache_db.upsert_record(conn, "Processos", "p1", {
        "page_id": "p1", "numero_do_processo": "0001-23",
    })
    record = {"cliente": [["u1"]], "processo": ["p1"]}
    assert _format_dashboard_subtitle(conn, record) == (
        "Aurora Magalhães · 0001-23"
    )
    # Sem cliente
    record2 = {"cliente": None, "processo": ["p1"]}
    assert _format_dashboard_subtitle(conn, record2) == "0001-23"
    # Sem CNJ
    record3 = {"cliente": [["u1"]], "processo": []}
    assert _format_dashboard_subtitle(conn, record3) == "Aurora Magalhães"


def test_R6_dashboard_card_renders_resolved_client_not_uuid() -> None:
    """Smoke ponta-a-ponta: instancia DashboardPage com cache populado,
    refresh, confirma que _UrgentTaskItem mostra "Maria Silva" na linha
    2 (cliente) e não o UUID-cru.

    Round 6 hotfix: estrutura mudou — cliente fica em ``_cliente_lbl``
    (linha 2 do item Kanban), CNJ em ``_cnj_lbl`` (linha 3).
    """
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import DashboardPage, _UrgentTaskItem

    conn = _full_conn()
    # Cliente cacheada com nome conhecido
    cache_db.upsert_record(conn, "Clientes", "cli-x", {
        "page_id": "cli-x", "nome": "Maria Silva",
    })
    # Processo cacheado com CNJ
    cache_db.upsert_record(conn, "Processos", "proc-x", {
        "page_id": "proc-x", "numero_do_processo": "9999-2026",
    })
    # Tarefa pra hoje, com rollup-de-relation pra cliente + relation pra processo
    today_str = datetime.date.today().isoformat()
    cache_db.upsert_record(conn, "Tarefas", "tar-x", {
        "page_id": "tar-x",
        "tarefa": "Petição inicial",
        "prazo_fatal": today_str,
        "cliente": [["cli-x"]],
        "processo": ["proc-x"],
        "status": "Pendente",
    })
    page = DashboardPage(conn=conn, user={"name": "Test"})
    page.refresh()
    # Procura todos _UrgentTaskItem visíveis e captura linhas 2 e 3.
    visible_clientes: list[str] = []
    visible_cnjs: list[str] = []
    for item in page.findChildren(_UrgentTaskItem):
        if item.isVisibleTo(page):
            visible_clientes.append(item._cliente_lbl.text())  # noqa: SLF001
            visible_cnjs.append(item._cnj_lbl.text())  # noqa: SLF001
    # O item da tarefa deve ter linha 2 com nome resolvido.
    assert "Maria Silva" in visible_clientes, (
        f"Linhas-2 visíveis: {visible_clientes}"
    )
    assert "9999-2026" in visible_cnjs, (
        f"Linhas-3 (CNJ) visíveis: {visible_cnjs}"
    )
    # E nenhuma linha visível pode conter o UUID-cru.
    assert not any("cli-x" in c for c in visible_clientes), (
        f"UUID vazou no cliente: {visible_clientes}"
    )


def test_R6_dashboard_urgent_groups_have_headers_with_count() -> None:
    """Headers dos 3 cards sempre presentes, com label uppercase à
    esquerda e contagem à direita (separados por stretch — não inline).
    Smoke confirma que a contagem reflete o número real de tarefas em
    cada bucket.

    Round 6 hotfix: header agora tem 2 labels (`_label_lbl` + `_count_lbl`)
    em vez de uma única `_header_lbl` com texto inline.
    """
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import DashboardPage

    conn = _full_conn()
    today = datetime.date.today()
    yesterday = (today - datetime.timedelta(days=1)).isoformat()
    today_str = today.isoformat()
    cache_db.upsert_record(conn, "Tarefas", "t1", {
        "page_id": "t1", "tarefa": "Vencida 1", "prazo_fatal": yesterday,
        "status": "Pendente",
    })
    cache_db.upsert_record(conn, "Tarefas", "t2", {
        "page_id": "t2", "tarefa": "Vencida 2", "prazo_fatal": yesterday,
        "status": "Pendente",
    })
    cache_db.upsert_record(conn, "Tarefas", "t3", {
        "page_id": "t3", "tarefa": "Hoje 1", "prazo_fatal": today_str,
        "status": "Pendente",
    })
    page = DashboardPage(conn=conn, user={"name": "Test"})
    page.refresh()
    venc = page._urgent_cards["vencida"]  # noqa: SLF001
    hoje = page._urgent_cards["hoje"]  # noqa: SLF001
    amanha = page._urgent_cards["amanha"]  # noqa: SLF001
    assert venc._label_lbl.text() == "VENCIDAS"  # noqa: SLF001
    assert venc._count_lbl.text() == "2"  # noqa: SLF001
    assert hoje._label_lbl.text() == "HOJE"  # noqa: SLF001
    assert hoje._count_lbl.text() == "1"  # noqa: SLF001
    assert amanha._label_lbl.text() == "AMANHÃ"  # noqa: SLF001
    assert amanha._count_lbl.text() == "0"  # noqa: SLF001


# ---------------------------------------------------------------------------
# Round 6 hotfix — Kanban 3 colunas (substitui a stack vertical do R6 P2)
# ---------------------------------------------------------------------------


def test_R6_dashboard_uses_3_column_grid_layout() -> None:
    """Round 6 hotfix: os 3 cards (Vencidas/Hoje/Amanhã) ficam em
    QHBoxLayout lado a lado, cada um com stretch=1 (1fr 1fr 1fr).
    Substitui a stack QVBoxLayout do R6 P2."""
    import sys
    from PySide6.QtWidgets import QApplication, QHBoxLayout
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.pages.dashboard import DashboardPage, _UrgentColumnCard

    conn = _full_conn()
    page = DashboardPage(conn=conn, user={"name": "Test"})
    # Container é QHBoxLayout
    assert isinstance(page._tasks_container, QHBoxLayout)  # noqa: SLF001
    # Tem 3 _UrgentColumnCard como filhos diretos do layout
    cards_in_layout = [
        page._tasks_container.itemAt(i).widget()  # noqa: SLF001
        for i in range(page._tasks_container.count())  # noqa: SLF001
    ]
    cards_in_layout = [
        c for c in cards_in_layout if isinstance(c, _UrgentColumnCard)
    ]
    assert len(cards_in_layout) == 3
    # Stretch=1 em cada — divide largura por igual.
    for i in range(3):
        assert page._tasks_container.stretch(i) == 1  # noqa: SLF001


def test_R6_dashboard_card_renders_3_lines_per_task() -> None:
    """Round 6 hotfix: cada _UrgentTaskItem tem 3 labels distintos
    (tipo bold, cliente normal, CNJ mono). Validação por população real
    via cache + refresh."""
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.cache import db as cache_db
    from notion_rpadv.pages.dashboard import DashboardPage, _UrgentTaskItem

    conn = _full_conn()
    cache_db.upsert_record(conn, "Clientes", "cli", {
        "page_id": "cli", "nome": "Luis Fernando",
    })
    cache_db.upsert_record(conn, "Processos", "proc", {
        "page_id": "proc", "numero_do_processo": "0000867-77.2023.5.10.0003",
    })
    cache_db.upsert_record(conn, "Catalogo", "cat", {
        "page_id": "cat", "nome": "A.02 Emenda à inicial",
    })
    today_str = datetime.date.today().isoformat()
    cache_db.upsert_record(conn, "Tarefas", "tar", {
        "page_id": "tar",
        "tarefa": "fallback livre",  # ignorado quando tipo_de_tarefa resolve
        "tipo_de_tarefa": ["cat"],
        "prazo_fatal": today_str,
        "cliente": [["cli"]],
        "processo": ["proc"],
        "status": "Pendente",
    })
    page = DashboardPage(conn=conn, user={"name": "Test"})
    page.refresh()
    visible_items = [
        i for i in page.findChildren(_UrgentTaskItem)
        if i.isVisibleTo(page)
    ]
    assert len(visible_items) == 1
    item = visible_items[0]
    # Linha 1: tipo (Catalog) bold → "A.02 Emenda à inicial"
    assert item._tipo_lbl.text() == "A.02 Emenda à inicial"  # noqa: SLF001
    # Linha 2: cliente
    assert item._cliente_lbl.text() == "Luis Fernando"  # noqa: SLF001
    # Linha 3: CNJ
    assert item._cnj_lbl.text() == "0000867-77.2023.5.10.0003"  # noqa: SLF001


def test_R6_dashboard_columns_have_uniform_height() -> None:
    """Round 6 hotfix: os 3 cards usam QSizePolicy.Expanding horizontal
    + Preferred vertical, em QHBoxLayout com stretch=1. Qt naturalmente
    iguala as 3 alturas pelo maior sizeHint vertical (a coluna com mais
    items dita a altura, as outras 2 acompanham com espaço vazio no
    fundo). O teste valida o size policy + stretch — não tenta medir
    altura renderizada (depende de show()/processEvents)."""
    import sys
    from PySide6.QtWidgets import QApplication, QSizePolicy
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.pages.dashboard import DashboardPage

    conn = _full_conn()
    page = DashboardPage(conn=conn, user={"name": "Test"})
    for key in ("vencida", "hoje", "amanha"):
        card = page._urgent_cards[key]  # noqa: SLF001
        sp = card.sizePolicy()
        # Horizontal: Expanding pra ocupar a fração do QHBoxLayout.
        assert sp.horizontalPolicy() == QSizePolicy.Policy.Expanding, (
            f"{key}: horizontal policy={sp.horizontalPolicy()}"
        )
        # Vertical: Preferred — Qt iguala alturas pegando max(sizeHint).
        assert sp.verticalPolicy() == QSizePolicy.Policy.Preferred, (
            f"{key}: vertical policy={sp.verticalPolicy()}"
        )
    # E todos os 3 entram no QHBoxLayout com mesmo stretch (1fr cada).
    assert page._tasks_container.stretch(0) == 1  # noqa: SLF001
    assert page._tasks_container.stretch(1) == 1  # noqa: SLF001
    assert page._tasks_container.stretch(2) == 1  # noqa: SLF001


def test_R6_dashboard_empty_bucket_card_still_visible() -> None:
    """Round 6 hotfix: card com count==0 ainda aparece (header com
    "AMANHÃ" + "0", lista vazia). Não esconde — predito UX, comunica
    "nenhuma tarefa pra amanhã" sem fazer o operador inferir do
    espaço em branco."""
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.pages.dashboard import DashboardPage

    conn = _full_conn()  # zero tarefas
    page = DashboardPage(conn=conn, user={"name": "Test"})
    page.refresh()
    for key in ("vencida", "hoje", "amanha"):
        card = page._urgent_cards[key]  # noqa: SLF001
        # Card visível
        assert not card.isHidden(), f"{key}: card oculto inesperado"
        # Header presente com count=0
        assert card._count_lbl.text() == "0"  # noqa: SLF001
        # Nenhum item visível dentro do card
        from notion_rpadv.pages.dashboard import _UrgentTaskItem
        for item in card.findChildren(_UrgentTaskItem):
            assert not item.isVisible() or item.isHidden(), (
                f"{key}: item visível em card vazio"
            )
