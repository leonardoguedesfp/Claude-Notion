"""Round 5 (30-abr-2026) — testes de rendering: people resolution + URL hyperlink.

Cobre:
- ``resolve_user_name`` (helper compartilhado entre model e exporter)
- ``_display_value`` branch ``people`` no BaseTableModel
- ``_format_for_excel`` branch ``url`` (Round 5 item 2)
- ``_write_base_sheet`` hyperlink em url no xlsx
"""
from __future__ import annotations

from unittest.mock import MagicMock

import pytest


# ---------------------------------------------------------------------------
# Item 3 — resolve_user_name helper + people branch in _display_value
# ---------------------------------------------------------------------------


def test_R5_resolve_user_name_known_uuid_returns_name() -> None:
    """UUID conhecido em NOTION_USERS retorna o nome configurado."""
    from notion_bulk_edit.config import resolve_user_name
    # UUID real da Déborah no NOTION_USERS
    name = resolve_user_name("23fd872b-594c-8178-840c-00029746e827")
    assert name == "Déborah"


def test_R5_resolve_user_name_unknown_uuid_falls_back_to_uuid() -> None:
    """UUID não em NOTION_USERS retorna o próprio UUID (bot, ex-membro,
    placeholder)."""
    from notion_bulk_edit.config import resolve_user_name
    assert resolve_user_name("uuid-fora-do-time") == "uuid-fora-do-time"


def test_R5_resolve_user_name_empty_returns_empty_string() -> None:
    """UUID vazio/None vira string vazia (caller normalmente filtra antes,
    mas helper é defensivo)."""
    from notion_bulk_edit.config import resolve_user_name
    assert resolve_user_name("") == ""


def test_R5_resolve_user_name_accepts_injected_users_dict() -> None:
    """``users=`` permite injeção pra testes sem depender do singleton
    (pattern já usado em snapshot_exporter)."""
    from notion_bulk_edit.config import resolve_user_name
    custom = {"u1": {"name": "Alice", "initials": "A"}}
    assert resolve_user_name("u1", users=custom) == "Alice"
    assert resolve_user_name("u2", users=custom) == "u2"


def test_R5_resolve_user_name_returns_uuid_when_name_field_missing() -> None:
    """Se a entrada existe mas não tem 'name', cai no UUID (defesa contra
    NOTION_USERS malformado)."""
    from notion_bulk_edit.config import resolve_user_name
    custom = {"u1": {"initials": "A"}}  # sem 'name'
    assert resolve_user_name("u1", users=custom) == "u1"


def test_R5_display_value_people_resolves_uuids_to_names() -> None:
    """``_display_value`` branch ``people`` resolve list de UUIDs pra
    nomes via NOTION_USERS (smoke do item 3 do Round 5)."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Responsável", tipo="people", label="Responsável",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    raw = ["23fd872b-594c-8178-840c-00029746e827"]
    assert _display_value(spec, raw) == "Déborah"


def test_R5_display_value_people_joins_multiple_with_comma() -> None:
    """Múltiplos UUIDs viram comma-separated."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Responsável", tipo="people", label="Responsável",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    raw = [
        "23fd872b-594c-8178-840c-00029746e827",  # Déborah
        "240d872b-594c-81f4-82e1-000212a926fc",  # Leonardo
    ]
    assert _display_value(spec, raw) == "Déborah, Leonardo"


def test_R5_display_value_people_falls_back_to_uuid_for_unknown() -> None:
    """UUID desconhecido vira ele mesmo (não some — visibilidade
    diagnóstica)."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Responsável", tipo="people", label="Responsável",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    raw = ["uuid-fora", "23fd872b-594c-8178-840c-00029746e827"]
    assert _display_value(spec, raw) == "uuid-fora, Déborah"


def test_R5_display_value_people_empty_list_returns_placeholder() -> None:
    """Lista vazia cai no placeholder em-dash do model (linhas
    137-140 do _display_value)."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Responsável", tipo="people", label="Responsável",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    assert _display_value(spec, []) == "—"


def test_R5_snapshot_exporter_people_uses_helper(tmp_path) -> None:
    """``_format_for_excel`` people branch agora usa resolve_user_name
    em vez de inline lookup. Comportamento ponta-a-ponta inalterado."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    notion_users = {"u1": {"name": "Alice"}}
    val, miss = _format_for_excel(["u1", "u2"], "people", {}, notion_users)
    assert val == "Alice, u2"
    assert miss == 0


# ---------------------------------------------------------------------------
# Item 2 — URL "link"/"indisponível" + hyperlink no xlsx
# ---------------------------------------------------------------------------


def test_R5_display_value_url_with_value_returns_link() -> None:
    """URL preenchido vira "link" no DisplayRole."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Link externo", tipo="url", label="Link externo",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    assert _display_value(spec, "https://example.com/proc/123") == "link"


def test_R5_display_value_url_empty_returns_indisponivel() -> None:
    """URL vazio (string em branco) vira "indisponível"."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Link externo", tipo="url", label="Link externo",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    assert _display_value(spec, "") == "indisponível"
    assert _display_value(spec, "   ") == "indisponível"


def test_R5_display_value_url_none_returns_indisponivel() -> None:
    """URL None vira "indisponível" (não em-dash placeholder)."""
    from notion_bulk_edit.schemas import PropSpec
    from notion_rpadv.models.base_table_model import _display_value
    spec = PropSpec(
        notion_name="Link externo", tipo="url", label="Link externo",
        editavel=True, obrigatorio=False, opcoes=(),
    )
    assert _display_value(spec, None) == "indisponível"


def test_R5_format_for_excel_url_with_value_returns_link() -> None:
    """``_format_for_excel`` url com URL retorna ("link", 0). Hyperlink
    real é aplicado pelo ``_write_base_sheet`` via cell.hyperlink."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(
        "https://lex.com/p/1", "url", {}, {},
    )
    assert val == "link"
    assert miss == 0


def test_R5_format_for_excel_url_empty_returns_indisponivel() -> None:
    """``_format_for_excel`` url vazio retorna ("indisponível", 0)."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    assert _format_for_excel(None, "url", {}, {}) == ("indisponível", 0)
    assert _format_for_excel("", "url", {}, {}) == ("indisponível", 0)
    assert _format_for_excel("   ", "url", {}, {}) == ("indisponível", 0)


def test_R5_xlsx_url_cell_has_hyperlink_when_value_present(tmp_path) -> None:
    """Smoke ponta-a-ponta: célula de url com URL tem cell.hyperlink
    apontando pro URL bruto, e display value 'link'.

    Usa base sintética ``Synth`` (fora de DEFAULT_LAYOUTS) pra que a
    ordem seja schema-driven e a célula de Link externo fique em col 2.
    """
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    pages = [{
        "id": "p1",
        "properties": {
            "Nome": {
                "type": "title",
                "title": [{"plain_text": "Processo X"}],
            },
            "Link externo": {
                "type": "url",
                "url": "https://lex.com/proc/abc-123",
            },
        },
    }]
    client = MagicMock()
    client.query_all.return_value = pages
    schemas = {"Synth": {"properties": {
        "nome": {
            "notion_name": "Nome", "tipo": "title", "label": "Nome",
            "default_order": 1, "default_visible": True, "opcoes": [],
        },
        "link_externo": {
            "notion_name": "Link externo", "tipo": "url",
            "label": "Link externo", "default_order": 2,
            "default_visible": True, "opcoes": [],
        },
    }}}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Synth"], dest_path=dest,
        schema_registry=reg, data_sources={"Synth": "ds-s"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Synth"]
    cell = ws.cell(row=2, column=2)  # Link externo é a 2ª coluna
    assert cell.value == "link"
    assert cell.hyperlink is not None
    # cell.hyperlink pode ser obj Hyperlink ou string dependendo de leitura;
    # checa o target em ambos os formatos.
    target = getattr(cell.hyperlink, "target", None) or str(cell.hyperlink)
    assert "lex.com/proc/abc-123" in target


def test_R5_xlsx_url_cell_indisponivel_has_no_hyperlink(tmp_path) -> None:
    """URL vazio: célula tem texto 'indisponível' como texto plano, sem
    hyperlink anexado."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    pages = [{
        "id": "p1",
        "properties": {
            "Nome": {
                "type": "title",
                "title": [{"plain_text": "Processo Y"}],
            },
            "Link externo": {"type": "url", "url": None},
        },
    }]
    client = MagicMock()
    client.query_all.return_value = pages
    schemas = {"Synth": {"properties": {
        "nome": {
            "notion_name": "Nome", "tipo": "title", "label": "Nome",
            "default_order": 1, "default_visible": True, "opcoes": [],
        },
        "link_externo": {
            "notion_name": "Link externo", "tipo": "url",
            "label": "Link externo", "default_order": 2,
            "default_visible": True, "opcoes": [],
        },
    }}}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Synth"], dest_path=dest,
        schema_registry=reg, data_sources={"Synth": "ds-s"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Synth"]
    cell = ws.cell(row=2, column=2)
    assert cell.value == "indisponível"
    assert cell.hyperlink is None


def test_R5_double_click_on_url_cell_opens_browser() -> None:
    """Round 5 item 2: double-click numa célula de url chama
    QDesktopServices.openUrl com o URL bruto. Mock captura sem abrir
    browser de verdade."""
    import sys
    from unittest.mock import patch
    from PySide6.QtCore import QModelIndex, Qt
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)

    # Mock simples de _on_table_double_clicked path: index com tipo url,
    # raw URL no EditRole. Não precisamos da page completa — só o método
    # com proxy/model substituídos por mocks.
    from notion_rpadv.pages.base_table_page import BaseTablePage

    page = MagicMock(spec=BaseTablePage)
    page._base = "Processos"

    # Fake spec do tipo url
    fake_spec = MagicMock()
    fake_spec.tipo = "url"
    fake_spec.target_base = ""

    # Fake model retorna URL no EditRole
    fake_model = MagicMock()
    fake_model.cols.return_value = ["link_externo"]
    fake_model.data.return_value = "https://lex.com/x"
    page._model = fake_model
    # Proxy passa-through
    page._proxy = MagicMock()
    page._proxy.mapToSource.side_effect = lambda i: i

    fake_index = MagicMock()
    fake_index.isValid.return_value = True
    fake_index.column.return_value = 0

    with patch(
        "notion_bulk_edit.schemas.get_prop", return_value=fake_spec,
    ), patch(
        "PySide6.QtGui.QDesktopServices.openUrl",
    ) as mock_open:
        BaseTablePage._on_table_double_clicked(page, fake_index)

    assert mock_open.called
    qurl = mock_open.call_args.args[0]
    assert qurl.toString() == "https://lex.com/x"


def test_R5_double_click_on_empty_url_cell_does_not_open_browser() -> None:
    """URL vazia → no-op, sem abrir browser."""
    import sys
    from unittest.mock import patch
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)

    from notion_rpadv.pages.base_table_page import BaseTablePage
    page = MagicMock(spec=BaseTablePage)
    page._base = "Processos"

    fake_spec = MagicMock()
    fake_spec.tipo = "url"
    fake_spec.target_base = ""

    fake_model = MagicMock()
    fake_model.cols.return_value = ["link_externo"]
    fake_model.data.return_value = None
    page._model = fake_model
    page._proxy = MagicMock()
    page._proxy.mapToSource.side_effect = lambda i: i

    fake_index = MagicMock()
    fake_index.isValid.return_value = True
    fake_index.column.return_value = 0

    with patch(
        "notion_bulk_edit.schemas.get_prop", return_value=fake_spec,
    ), patch(
        "PySide6.QtGui.QDesktopServices.openUrl",
    ) as mock_open:
        BaseTablePage._on_table_double_clicked(page, fake_index)

    assert not mock_open.called


# ---------------------------------------------------------------------------
# Item 5 — xlsx escreve colunas na ordem do layout (visíveis, depois ocultas)
# ---------------------------------------------------------------------------


def _schema_with_props(props: list[tuple[str, str, str]]) -> dict:
    """Helper local: schema com (slug, notion_name, tipo) por entrada,
    preservando ordem de inserção como default_order."""
    props_dict = {}
    for order, (slug, notion_name, tipo) in enumerate(props, start=1):
        props_dict[slug] = {
            "notion_name": notion_name,
            "tipo": tipo,
            "label": notion_name,
            "default_order": order,
            "default_visible": True,
            "opcoes": [],
        }
    return {"properties": props_dict}


def test_R5_xlsx_columns_in_layout_order_visible_first(tmp_path) -> None:
    """Round 5 item 5: xlsx escreve colunas na ordem do layout (visíveis
    primeiro na ordem editorial), depois as omitidas (na ordem do schema).

    Catalogo é a base mais simples pra testar isso — layout tem 3
    visíveis (nome, categoria, observacoes) e o schema real tem mais 4
    (prazo, tarefas, criado_em, atualizado_em) que aparecem ocultas
    no app mas EXPORTAM TODAS no xlsx.
    """
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    # Schema sintético de Catalogo: ordem da API embaralhada de propósito
    # (categoria antes de nome, prazo no meio) pra provar que o exporter
    # reordena conforme o layout, não conforme o schema.
    schemas = {"Catalogo": _schema_with_props([
        ("categoria",     "Categoria",     "select"),
        ("prazo",         "Prazo",         "number"),
        ("nome",          "Nome",          "title"),
        ("tarefas",       "Tarefas",       "relation"),
        ("observacoes",   "Observações",   "rich_text"),
        ("criado_em",     "Criado em",     "created_time"),
        ("atualizado_em", "Atualizado em", "last_edited_time"),
    ])}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Catalogo"], dest_path=dest,
        schema_registry=reg, data_sources={"Catalogo": "ds-cat"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Catalogo"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, 8)
    ]
    # Visíveis no layout (Catalogo): nome, categoria, observacoes
    # Ocultas (resto do schema na ordem da API):
    #   categoria (já visível, skip), prazo, tarefas, criado_em,
    #   atualizado_em → excluindo as visíveis: prazo, tarefas,
    #   criado_em, atualizado_em.
    assert headers[:3] == ["Nome", "Categoria", "Observações"]
    # Próximas devem ser as ocultas, na ordem do schema (que aqui foi
    # embaralhado: prazo aparece antes de tarefas no schema).
    assert headers[3:] == [
        "Prazo", "Tarefas", "Criado em", "Atualizado em",
    ]


def test_R5_xlsx_columns_for_unknown_base_falls_back_to_schema_order(
    tmp_path,
) -> None:
    """Bases não cobertas pelo DEFAULT_LAYOUTS preservam a ordem do
    schema (compatibilidade com bases novas adicionadas sem layout
    editorial)."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    schemas = {"BaseSemLayout": _schema_with_props([
        ("alpha",   "Alpha",   "title"),
        ("delta",   "Delta",   "select"),
        ("bravo",   "Bravo",   "number"),
        ("charlie", "Charlie", "rich_text"),
    ])}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["BaseSemLayout"], dest_path=dest,
        schema_registry=reg, data_sources={"BaseSemLayout": "ds-x"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["BaseSemLayout"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    # Ordem da API preservada (sem layout pra reordenar).
    assert headers == ["Alpha", "Delta", "Bravo", "Charlie"]


def test_R5_xlsx_exports_all_schema_columns_even_when_reordered(
    tmp_path,
) -> None:
    """Round 5 item 5 garantia: ainda que reordenado, NENHUMA coluna do
    schema é perdida no xlsx (ordem muda, total não)."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    # Schema com 5 props; layout de Tarefas tem 10 slugs visíveis (mas
    # só responsavel está nesta intersecção mínima).
    schemas = {"Tarefas": _schema_with_props([
        ("status",       "Status",      "select"),
        ("descricao",    "Descrição",   "rich_text"),
        ("tarefa",       "Tarefa",      "title"),
        ("responsavel",  "Responsável", "people"),
        ("criado_em",    "Criado em",   "created_time"),
    ])}
    reg = MagicMock()
    reg._schemas = schemas
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Tarefas"], dest_path=dest,
        schema_registry=reg, data_sources={"Tarefas": "ds-t"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Tarefas"]
    # Coleta todos os headers até a primeira célula vazia.
    headers: list = []
    c = 1
    while True:
        v = ws.cell(row=1, column=c).value
        if v is None:
            break
        headers.append(v)
        c += 1
    # Sem perdas: todas as 5 props do schema presentes.
    assert set(headers) == {
        "Status", "Descrição", "Tarefa", "Responsável", "Criado em",
    }
    assert len(headers) == 5
