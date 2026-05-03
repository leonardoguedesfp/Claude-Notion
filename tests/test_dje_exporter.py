"""Testes do ``notion_rpadv.services.dje_exporter``: nome do arquivo,
versionamento, ordem de colunas, serialização JSON de arrays, conteúdo
da primeira linha."""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path


# ---------------------------------------------------------------------------
# Naming + versionamento
# ---------------------------------------------------------------------------


def test_format_filename_dd_mm_aa() -> None:
    """``Publicacoes_DJEN_{dd.mm.aa}_a_{dd.mm.aa}_v{N}.xlsx``."""
    from notion_rpadv.services.dje_exporter import format_filename
    name = format_filename(date(2026, 5, 1), date(2026, 5, 3), 1)
    assert name == "Publicacoes_DJEN_01.05.26_a_03.05.26_v1.xlsx"


def test_format_filename_intervalo_de_um_dia() -> None:
    """Quando intervalo = 1 dia (inicio == fim), formato preserva."""
    from notion_rpadv.services.dje_exporter import format_filename
    name = format_filename(date(2026, 5, 1), date(2026, 5, 1), 2)
    assert name == "Publicacoes_DJEN_01.05.26_a_01.05.26_v2.xlsx"


def test_next_version_v1_quando_vazio(tmp_path: Path) -> None:
    """Diretório vazio → próxima versão é 1."""
    from notion_rpadv.services.dje_exporter import next_version
    v = next_version(tmp_path, date(2026, 5, 1), date(2026, 5, 1))
    assert v == 1


def test_next_version_pula_v1_existente(tmp_path: Path) -> None:
    """v1 já existe → próximo é v2 (Round 7 case 6 do spec)."""
    from notion_rpadv.services.dje_exporter import (
        format_filename,
        next_version,
    )
    di, df = date(2026, 5, 1), date(2026, 5, 1)
    (tmp_path / format_filename(di, df, 1)).write_text("")
    v = next_version(tmp_path, di, df)
    assert v == 2


def test_next_version_pula_ate_achar_livre(tmp_path: Path) -> None:
    """v1, v2, v3 existem → v4."""
    from notion_rpadv.services.dje_exporter import (
        format_filename,
        next_version,
    )
    di, df = date(2026, 5, 1), date(2026, 5, 1)
    for v in (1, 2, 3):
        (tmp_path / format_filename(di, df, v)).write_text("")
    assert next_version(tmp_path, di, df) == 4


def test_next_version_diferente_intervalo_nao_colide(tmp_path: Path) -> None:
    """v1 do intervalo X não interfere com v1 do intervalo Y."""
    from notion_rpadv.services.dje_exporter import (
        format_filename,
        next_version,
    )
    other_di, other_df = date(2026, 4, 30), date(2026, 4, 30)
    (tmp_path / format_filename(other_di, other_df, 1)).write_text("")
    # Intervalo diferente → ainda v1.
    di, df = date(2026, 5, 1), date(2026, 5, 1)
    assert next_version(tmp_path, di, df) == 1


# ---------------------------------------------------------------------------
# write_publicacoes_xlsx — content tests (Round 7 case 5)
# ---------------------------------------------------------------------------


def _read_back(path: Path):
    from openpyxl import load_workbook
    return load_workbook(path)


def test_xlsx_aba_chamada_publicacoes(tmp_path: Path) -> None:
    """Aba única se chama 'Publicacoes' (sem acento — coerência com
    o nome do arquivo)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    result = write_publicacoes_xlsx(
        rows=[],
        output_dir=tmp_path,
        data_inicio=date(2026, 5, 1),
        data_fim=date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    assert wb.sheetnames == ["Publicacoes"]


def test_xlsx_datadisponibilizacao_eh_primeira_coluna(tmp_path: Path) -> None:
    """Pós-Fase 3: ``datadisponibilizacao`` (formato BR DD/MM/AAAA) é a
    1ª coluna do schema canônico (visível). ``advogados_consultados_escritorio``
    permanece exportada mas oculta (column_dimensions.hidden=True)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    rows = [{
        "id": 1, "hash": "abc", "texto": "publi 1",
        "datadisponibilizacao": "01/05/2026",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [
            {"advogado": {"numero_oab": "36129", "uf_oab": "DF", "nome": "Leonardo"}},
        ],
    }]
    result = write_publicacoes_xlsx(
        rows=rows,
        output_dir=tmp_path,
        data_inicio=date(2026, 5, 1),
        data_fim=date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=1, column=1).value == "datadisponibilizacao"
    assert ws.cell(row=2, column=1).value == "01/05/2026"
    # advogados_consultados_escritorio: ainda está exportada, mas oculta.
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    advs_idx = headers.index("advogados_consultados_escritorio") + 1
    assert ws.cell(row=2, column=advs_idx).value == (
        "Leonardo Guedes da Fonseca Passos (36129/DF)"
    )


def test_xlsx_header_em_negrito(tmp_path: Path) -> None:
    """Headers em bold pra destacar."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    rows = [{"advogado_consultado": "X", "id": 1}]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 5, 1), date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=1, column=1).font.bold is True


def test_xlsx_serializa_arrays_como_json(tmp_path: Path) -> None:
    """Round 7 spec: arrays (destinatarios, destinatarioadvogados)
    viram JSON string com ensure_ascii=False (acentos preservados).
    F2: agora dentro do schema canônico (cols 18 e 19)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    rows = [{
        "advogado_consultado": "X",
        "id": 1,
        "destinatarios": [
            {"nome": "Luís Fernando", "polo": "ATIVO"},
            {"nome": "Mariana Souza", "polo": "PASSIVO"},
        ],
        "destinatarioadvogados": [
            {"numero_oab": "12345", "uf_oab": "DF", "nome": "X"},
        ],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 5, 1), date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    dest_idx = headers.index("destinatarios") + 1
    advs_idx = headers.index("destinatarioadvogados") + 1
    dest_value = ws.cell(row=2, column=dest_idx).value
    advs_value = ws.cell(row=2, column=advs_idx).value
    # JSON string parseável (não Python repr)
    parsed_dest = json.loads(dest_value)
    parsed_advs = json.loads(advs_value)
    assert parsed_dest[0]["nome"] == "Luís Fernando"
    assert parsed_advs[0]["numero_oab"] == "12345"
    # ensure_ascii=False — acentos literais (não \\u).
    assert "Luís" in dest_value


def test_xlsx_schema_canonico_20_colunas_em_ordem_fixa(tmp_path: Path) -> None:
    """Pós-Fase 3: schema canônico passou de 21 → 20 colunas em ordem
    fixa (CANONICAL_COLUMNS). Eliminada ``data_disponibilizacao`` (com
    underscore) — duplicata da ``datadisponibilizacao`` que vem da API.
    Nova ordem: 9 visíveis primeiro (datadisponibilizacao, siglaTribunal,
    ..., link), depois 11 ocultas (advogados_consultados_escritorio,
    observacoes, ..., oabs_externas_consultadas)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    from notion_rpadv.services.dje_transform import CANONICAL_COLUMNS
    rows = [{
        "id":            1,
        "hash":          "abc",
        "siglaTribunal": "TRT10",
        "tipoDocumento": "Intimação",
        "advogado_consultado": "X (1/DF)",
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 5, 1), date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    # 20 colunas exatas, na ordem do schema canônico pós-Fase 3.
    assert headers == CANONICAL_COLUMNS
    assert headers[0] == "datadisponibilizacao"
    assert headers[1] == "siglaTribunal"
    assert headers[-1] == "oabs_externas_consultadas"


def test_xlsx_chaves_extras_no_payload_sao_ignoradas(tmp_path: Path) -> None:
    """Round 7 Fase 2: schema canônico filtra chaves novas que possam
    aparecer no payload do DJEN — só as 20 do CANONICAL_COLUMNS entram
    no xlsx.

    Substitui o teste F1 ``test_xlsx_chaves_novas_em_items_subsequentes``
    que validava o oposto (auto-include de chaves novas) — comportamento
    desejado da F1 mas explícito como anti-padrão na F2 (schema é
    contrato).
    """
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    from notion_rpadv.services.dje_transform import CANONICAL_COLUMNS
    rows = [
        {"advogado_consultado": "X", "id": 1, "hash": "a", "extra_inesperado": "novo"},
    ]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 5, 1), date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    assert headers == CANONICAL_COLUMNS
    assert "extra_inesperado" not in headers


def test_xlsx_lista_vazia_gera_arquivo_so_com_header(tmp_path: Path) -> None:
    """0 publicações → arquivo gerado mesmo assim com header completo
    (20 colunas do schema canônico pós-Fase 3)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    from notion_rpadv.services.dje_transform import CANONICAL_COLUMNS
    result = write_publicacoes_xlsx(
        rows=[],
        output_dir=tmp_path,
        data_inicio=date(2026, 5, 1),
        data_fim=date(2026, 5, 1),
    )
    assert result.path.exists()
    assert result.skipped == []
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # Pós-Fase 3: 1ª col é datadisponibilizacao (visível).
    assert ws.cell(row=1, column=1).value == "datadisponibilizacao"
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    assert headers == CANONICAL_COLUMNS
    # Sem linhas de dados.
    assert ws.cell(row=2, column=1).value is None


def test_xlsx_versionamento_em_2_runs_consecutivas(tmp_path: Path) -> None:
    """Round 7 spec case 6: rodar 2× o mesmo intervalo → v2 no segundo."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    di, df = date(2026, 5, 1), date(2026, 5, 1)
    r1 = write_publicacoes_xlsx([], tmp_path, di, df)
    r2 = write_publicacoes_xlsx([], tmp_path, di, df)
    assert r1.path.name.endswith("_v1.xlsx")
    assert r2.path.name.endswith("_v2.xlsx")
    # Ambos sobrevivem (não sobrescreveu)
    assert r1.path.exists()
    assert r2.path.exists()


def test_xlsx_cria_diretorio_se_nao_existir(tmp_path: Path) -> None:
    """write_publicacoes_xlsx faz mkdir(parents=True) — caller pode
    passar caminho sob diretório que ainda não foi criado."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    target = tmp_path / "subdir1" / "subdir2"
    assert not target.exists()
    write_publicacoes_xlsx(
        [], target, date(2026, 5, 1), date(2026, 5, 1),
    )
    assert target.exists()


def test_xlsx_none_em_campo_canonico_vira_celula_vazia(tmp_path: Path) -> None:
    """Valor None em campo do schema canônico vira célula vazia (não
    string 'None'). F2: testa via campo ``link`` que está no schema
    canônico (col 17) — F1 testava com chave fora do schema (que agora
    seria filtrada pelo transform)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    rows = [{
        "advogado_consultado": "X",
        "id": 1,
        "link": None,  # campo canônico mas com valor None
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 5, 1), date(2026, 5, 1),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    idx = headers.index("link") + 1
    assert ws.cell(row=2, column=idx).value is None


# ---------------------------------------------------------------------------
# Fase 2.1 — exporter robusto a erro por linha (Bug B parte 2)
# ---------------------------------------------------------------------------


def test_F21_05_smoke_endtoend_com_u2426_em_texto_e_destinatarios(
    tmp_path: Path,
) -> None:
    """End-to-end: row contém U+2426 SYMBOL FOR SUBSTITUTE FORM TWO no
    campo ``texto`` E em ``destinatarios`` (lista de dicts) → arquivo
    é gerado SEM erro. Defesa primária (transform) limpa ``texto``;
    defesa secundária (exporter pós-JSON) limpa ``destinatarios``."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    sujo = chr(0x2426)
    rows = [{
        "id": 1, "hash": "a", "siglaTribunal": "STJ",
        "data_disponibilizacao": "2026-04-29",
        "advogado_consultado": "X (1/DF)",
        "texto": f"AREsp 2427258/DF {sujo} PREVI",
        "destinatarios": [
            {"nome": f"Polo {sujo} Ativo", "polo": "ATIVO"},
        ],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 29), date(2026, 4, 29),
    )
    # Arquivo gerado; nenhuma linha pulada.
    assert result.path.exists()
    assert result.skipped == []
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    texto_idx = headers.index("texto") + 1
    dest_idx = headers.index("destinatarios") + 1
    texto_value = ws.cell(row=2, column=texto_idx).value
    dest_value = ws.cell(row=2, column=dest_idx).value
    # Caractere ilegal removido em ambas as camadas.
    assert sujo not in texto_value
    assert sujo not in dest_value
    # Conteúdo útil preservado.
    assert "AREsp 2427258/DF" in texto_value
    assert "Polo" in dest_value
    assert "Ativo" in dest_value


def test_F21_06_linha_com_erro_inesperado_eh_pulada_demais_escritas(
    tmp_path: Path, monkeypatch,
) -> None:
    """Mock força ``ws.cell(...)`` a levantar quando o VALOR contém um
    marker mágico → linha que tem esse marker em qualquer célula é
    pulada (entry em ``result.skipped``); demais linhas permanecem no
    arquivo (sem gaps).

    Mock baseado em valor (não em row index) é determinístico: como o
    exporter NÃO avança ``written_row_idx`` em falha, mockar por row
    index faria as linhas seguintes baterem no mesmo slot e
    aparentemente "também falharem"."""
    from openpyxl import Workbook as _OrigWB
    from openpyxl.utils.exceptions import IllegalCharacterError

    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    MAGIC = "FAIL_THIS_ROW"
    rows = [
        {"id": 100, "advogado_consultado": "A", "siglaTribunal": "X1",
         "data_disponibilizacao": "2026-04-29", "texto": "linha1 ok"},
        {"id": 200, "advogado_consultado": "B", "siglaTribunal": "X2",
         "data_disponibilizacao": "2026-04-29", "texto": f"linha2 {MAGIC}"},
        {"id": 300, "advogado_consultado": "C", "siglaTribunal": "X3",
         "data_disponibilizacao": "2026-04-29", "texto": "linha3 ok"},
    ]
    orig_active = _OrigWB.active.fget

    def _bad_active(self):
        ws = orig_active(self)
        original_cell = ws.cell

        def patched(row=None, column=None, value=None):
            if isinstance(value, str) and MAGIC in value:
                raise IllegalCharacterError("fake fail magic")
            return original_cell(row=row, column=column, value=value)

        ws.cell = patched
        return ws

    monkeypatch.setattr(_OrigWB, "active", property(_bad_active))

    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 29), date(2026, 4, 29),
    )
    # Desfaz patch antes de ler de volta (load_workbook usa Workbook.active
    # como setter, que nosso patch não fornece — undoes restaura o original).
    monkeypatch.undo()
    # Exatamente 1 linha pulada (a do MAGIC).
    assert len(result.skipped) == 1
    sk = result.skipped[0]
    # source_idx é 1-based no contexto pós-transform (deduped+sorted).
    # ID 200 corresponde ao MAGIC.
    assert sk.row_id == 200
    assert "fake fail" in sk.error.lower()
    # Arquivo existe e tem header + 2 linhas de dados (sem gaps).
    # Pós-Fase 3: col 2 = siglaTribunal (visível). É o campo mais
    # confiável pra checar presença de dado nessas rows sintéticas
    # (col 1 = datadisponibilizacao, populada via post-processing).
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=2, column=2).value is not None
    assert ws.cell(row=3, column=2).value is not None
    # Linha 4 do xlsx é vazia (só foram escritas 2 de 3 source rows).
    assert ws.cell(row=4, column=2).value is None


def test_F21_07_multiplas_linhas_com_erro_todas_puladas_arquivo_gerado(
    tmp_path: Path, monkeypatch, caplog,
) -> None:
    """Múltiplas linhas com marker mágico falham → todas puladas, log
    warning emitido pra cada, arquivo gerado com as restantes."""
    import logging as _logging

    from openpyxl import Workbook as _OrigWB
    from openpyxl.utils.exceptions import IllegalCharacterError

    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    MAGIC = "FAIL_THIS_ROW"
    # 5 rows; ids 200 e 400 contêm MAGIC → falham.
    rows = [
        {"id": 100, "advogado_consultado": "A0", "siglaTribunal": "T0",
         "data_disponibilizacao": "2026-04-29", "texto": "ok 0"},
        {"id": 200, "advogado_consultado": "A1", "siglaTribunal": "T1",
         "data_disponibilizacao": "2026-04-29", "texto": f"fail {MAGIC} 1"},
        {"id": 300, "advogado_consultado": "A2", "siglaTribunal": "T2",
         "data_disponibilizacao": "2026-04-29", "texto": "ok 2"},
        {"id": 400, "advogado_consultado": "A3", "siglaTribunal": "T3",
         "data_disponibilizacao": "2026-04-29", "texto": f"fail {MAGIC} 3"},
        {"id": 500, "advogado_consultado": "A4", "siglaTribunal": "T4",
         "data_disponibilizacao": "2026-04-29", "texto": "ok 4"},
    ]
    orig_active = _OrigWB.active.fget

    def _bad_active(self):
        ws = orig_active(self)
        original_cell = ws.cell

        def patched(row=None, column=None, value=None):
            if isinstance(value, str) and MAGIC in value:
                raise IllegalCharacterError("fake fail magic")
            return original_cell(row=row, column=column, value=value)

        ws.cell = patched
        return ws

    monkeypatch.setattr(_OrigWB, "active", property(_bad_active))

    with caplog.at_level(_logging.WARNING, logger="dje.exporter"):
        result = write_publicacoes_xlsx(
            rows, tmp_path, date(2026, 4, 29), date(2026, 4, 29),
        )
    # Undo do patch antes de _read_back (load_workbook precisa do setter).
    monkeypatch.undo()
    # 2 linhas puladas (ids 200 e 400).
    assert len(result.skipped) == 2
    skipped_ids = {sk.row_id for sk in result.skipped}
    assert skipped_ids == {200, 400}
    # Log warning emitido pra cada uma (mensagem contém "linha pulada").
    warn_msgs = [
        r.getMessage() for r in caplog.records
        if r.levelno == _logging.WARNING and r.name == "dje.exporter"
    ]
    pulada_count = sum("linha pulada" in m for m in warn_msgs)
    assert pulada_count == 2, (
        f"esperava 2 warnings 'linha pulada', got {pulada_count}: {warn_msgs!r}"
    )
    # Arquivo gerado com 3 linhas de dados (5 source - 2 puladas).
    # Pós-Fase 3: col 2 = siglaTribunal (visível, sempre populada
    # nessas rows sintéticas).
    assert result.path.exists()
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=2, column=2).value is not None
    assert ws.cell(row=3, column=2).value is not None
    assert ws.cell(row=4, column=2).value is not None
    assert ws.cell(row=5, column=2).value is None


def test_F21_08_zero_linhas_com_erro_skipped_eh_lista_vazia(
    tmp_path: Path,
) -> None:
    """Sem nenhum erro de escrita → ``result.skipped == []`` (lista
    vazia, não None) e comportamento idêntico ao F2 anterior."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    rows = [{
        "id": 1, "advogado_consultado": "X (1/DF)",
        "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-29",
        "texto": "publicação normal sem lixo",
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 29), date(2026, 4, 29),
    )
    assert result.skipped == []
    assert isinstance(result.skipped, list)
    # Arquivo OK com 1 linha de dados. Pós-Fase 3: col 2 = siglaTribunal
    # (visível, sempre populada nessa row sintética).
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=2, column=2).value is not None
    assert ws.cell(row=3, column=2).value is None


# ---------------------------------------------------------------------------
# Fase 3 — F3-29..F3-31: schema renomeado e nova coluna
# ---------------------------------------------------------------------------


def test_F3_29_schema_xlsx_tem_20_colunas_em_ordem(tmp_path: Path) -> None:
    """Pós-Fase 3: o xlsx gerado tem 20 cabeçalhos exatamente nessa ordem
    (era 21 — ``data_disponibilizacao`` redundante eliminada)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    from notion_rpadv.services.dje_transform import CANONICAL_COLUMNS

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, 21)
    ]
    assert len(headers) == 20
    assert headers == CANONICAL_COLUMNS


def test_F3_30_visiveis_primeiro_advogados_escritorio_oculta(
    tmp_path: Path,
) -> None:
    """Pós-Fase 3: visíveis (datadisponibilizacao, siglaTribunal, ...) vêm
    primeiro; ``advogados_consultados_escritorio`` permanece exportada
    mas com ``column_dimensions.hidden=True`` (na fatia oculta a partir
    da col 10)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # 1ª col é datadisponibilizacao (visível).
    assert ws.cell(row=1, column=1).value == "datadisponibilizacao"
    # advogados_consultados_escritorio fica em algum lugar a partir da 10ª.
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    advs_idx = headers.index("advogados_consultados_escritorio") + 1
    assert advs_idx >= 10, (
        f"advogados_consultados_escritorio deveria estar na fatia oculta "
        f"(col >= 10), está na col {advs_idx}"
    )


def test_F3_31_oabs_externas_consultadas_eh_ultima_e_vazia_em_padrao(
    tmp_path: Path,
) -> None:
    """Pós-Fase 3: ``oabs_externas_consultadas`` é a 20ª coluna (oculta);
    vazia em modo padrão (defaults do ``write_publicacoes_xlsx``)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [
            {"advogado": {"numero_oab": "36129", "uf_oab": "DF", "nome": "Leonardo"}},
        ],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    assert ws.cell(row=1, column=20).value == "oabs_externas_consultadas"
    # Linha 2 (primeira de dados) tem essa célula vazia em modo padrão.
    assert ws.cell(row=2, column=20).value in (None, "")


# ---------------------------------------------------------------------------
# Fase 3 — F3-34..F3-37: histórico completo
# ---------------------------------------------------------------------------


def test_F3_34_historico_gerado_apos_execucao_com_publicacoes_novas(
    tmp_path: Path,
) -> None:
    """F3-34: após execução com publicações novas, gera/sobrescreve
    ``Historico_DJEN_completo.xlsx`` com todas as rows ordenadas."""
    from notion_rpadv.services.dje_exporter import (
        HISTORICO_FILENAME,
        write_historico_completo_xlsx,
    )

    db_rows = [
        {
            "id": 100, "hash": "h100", "siglaTribunal": "TRT10",
            "data_disponibilizacao": "2026-04-29",
            "destinatarioadvogados": [],
            "advogados_consultados_escritorio": "Ricardo (15523/DF)",
            "oabs_externas_consultadas": "",
        },
        {
            "id": 200, "hash": "h200", "siglaTribunal": "TRT01",
            "data_disponibilizacao": "2026-04-30",
            "destinatarioadvogados": [],
            "advogados_consultados_escritorio": "Leonardo (36129/DF)",
            "oabs_externas_consultadas": "",
        },
    ]
    result = write_historico_completo_xlsx(db_rows, tmp_path)
    assert result.locked is False
    assert result.path == tmp_path / HISTORICO_FILENAME
    assert result.path.exists()

    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # 1 header + 2 dados. Pós-Fase 3: 1ª col é datadisponibilizacao.
    assert ws.cell(row=1, column=1).value == "datadisponibilizacao"
    # Ordem: data DESC, sigla ASC
    # 200 (2026-04-30, TRT01) primeiro, 100 (2026-04-29, TRT10) depois.
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    id_col = headers.index("id") + 1
    assert ws.cell(row=2, column=id_col).value == 200
    assert ws.cell(row=3, column=id_col).value == 100


def test_F3_35_historico_inclui_publicacoes_de_modo_manual(
    tmp_path: Path,
) -> None:
    """F3-35: histórico junta publicações de execuções padrão e manual.
    Em particular, oabs_externas_consultadas aparece preenchida."""
    from notion_rpadv.services.dje_exporter import (
        write_historico_completo_xlsx,
    )

    db_rows = [
        {
            "id": 1, "hash": "h1", "siglaTribunal": "STJ",
            "data_disponibilizacao": "2026-04-15",
            "destinatarioadvogados": [],
            "advogados_consultados_escritorio": "Ricardo (15523/DF)",
            "oabs_externas_consultadas": "",
        },
        {
            "id": 2, "hash": "h2", "siglaTribunal": "STJ",
            "data_disponibilizacao": "2026-04-20",
            "destinatarioadvogados": [],
            "advogados_consultados_escritorio": "",
            "oabs_externas_consultadas": "Joao Silva (12345/SP)",
        },
    ]
    result = write_historico_completo_xlsx(db_rows, tmp_path)
    assert result.locked is False
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # 1ª linha de dados (mais recente, id=2) tem externa preenchida.
    # Pós-Fase 3: oabs_externas_consultadas é a col 20 (era 21).
    assert ws.cell(row=2, column=20).value == "Joao Silva (12345/SP)"
    # 2ª linha (id=1) tem externa vazia
    assert ws.cell(row=3, column=20).value in (None, "")


def test_F3_36_historico_locked_nao_derruba_execucao(tmp_path: Path) -> None:
    """F3-36: PermissionError ao escrever o .tmp → returna locked=True;
    Excel-de-execução foi gerado normalmente em outro path; SQLite OK."""
    from notion_rpadv.services.dje_exporter import (
        HISTORICO_TMP_FILENAME,
        write_historico_completo_xlsx,
    )
    import unittest.mock as mock

    db_rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "destinatarioadvogados": [],
        "advogados_consultados_escritorio": "Ricardo (15523/DF)",
        "oabs_externas_consultadas": "",
    }]
    # Simula bloqueio: patch Workbook.save pra levantar PermissionError
    # quando alvo é o .tmp.
    real_save = None
    from openpyxl.workbook.workbook import Workbook as _WB
    real_save = _WB.save

    def fake_save(self, target):
        if HISTORICO_TMP_FILENAME in str(target):
            raise PermissionError(
                f"[Errno 13] Permission denied: '{target}'",
            )
        return real_save(self, target)

    with mock.patch.object(_WB, "save", new=fake_save):
        result = write_historico_completo_xlsx(db_rows, tmp_path)
    assert result.locked is True
    assert result.path is None
    # Arquivo final NÃO existe (escrita pulada)
    from notion_rpadv.services.dje_exporter import HISTORICO_FILENAME
    assert not (tmp_path / HISTORICO_FILENAME).exists()


def test_F3_37_historico_em_banco_vazio_nao_eh_gerado_pelo_caller(
    tmp_path: Path,
) -> None:
    """F3-37 (parte exporter): chamar ``write_historico_completo_xlsx``
    com lista vazia gera arquivo só com header (caller deveria não chamar
    quando count=0; mas se chamar, ainda funciona sem crashar).

    O fluxo de não-gerar quando banco vazio vive no ``_DJEWorker._run_inner``
    (validado em test_leitor_dje_page com mock).
    """
    from notion_rpadv.services.dje_exporter import (
        write_historico_completo_xlsx,
    )

    result = write_historico_completo_xlsx([], tmp_path)
    assert result.locked is False
    assert result.path is not None
    assert result.path.exists()
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # Header existe; sem dados. Pós-Fase 3: 1ª col é datadisponibilizacao.
    assert ws.cell(row=1, column=1).value == "datadisponibilizacao"
    assert ws.cell(row=2, column=1).value is None


def test_F3_historico_atomic_rename_substitui_anterior(tmp_path: Path) -> None:
    """Quando histórico já existe, replace funciona e não deixa .tmp órfão."""
    from notion_rpadv.services.dje_exporter import (
        HISTORICO_FILENAME,
        HISTORICO_TMP_FILENAME,
        write_historico_completo_xlsx,
    )

    db_rows1 = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "destinatarioadvogados": [],
        "advogados_consultados_escritorio": "Ricardo (15523/DF)",
        "oabs_externas_consultadas": "",
    }]
    write_historico_completo_xlsx(db_rows1, tmp_path)
    final_path = tmp_path / HISTORICO_FILENAME
    assert final_path.exists()
    size1 = final_path.stat().st_size

    db_rows2 = db_rows1 + [{
        "id": 2, "hash": "h2", "siglaTribunal": "TRT01",
        "data_disponibilizacao": "2026-05-01",
        "destinatarioadvogados": [],
        "advogados_consultados_escritorio": "Leonardo (36129/DF)",
        "oabs_externas_consultadas": "",
    }]
    write_historico_completo_xlsx(db_rows2, tmp_path)
    assert final_path.exists()
    size2 = final_path.stat().st_size
    # Cresceu: 2ª escrita tem 1 linha a mais
    assert size2 != size1
    # .tmp não ficou órfão
    assert not (tmp_path / HISTORICO_TMP_FILENAME).exists()


# ---------------------------------------------------------------------------
# Pós-Fase 3 (2026-05-02) — reordenação + ocultação de colunas no xlsx
# ---------------------------------------------------------------------------


def test_pos_F3_visiveis_estao_em_ordem_no_inicio(tmp_path: Path) -> None:
    """Pós-Fase 3: as 9 colunas visíveis aparecem nas 9 primeiras
    posições, na ordem editorial: datadisponibilizacao, siglaTribunal,
    numeroprocessocommascara, nomeOrgao, tipoComunicacao, tipoDocumento,
    nomeClasse, texto, link."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers_visiveis = [
        ws.cell(row=1, column=c).value for c in range(1, 10)
    ]
    assert headers_visiveis == [
        "datadisponibilizacao",
        "siglaTribunal",
        "numeroprocessocommascara",
        "nomeOrgao",
        "tipoComunicacao",
        "tipoDocumento",
        "nomeClasse",
        "texto",
        "link",
    ]


def test_pos_F3_colunas_ocultas_tem_column_dimensions_hidden(
    tmp_path: Path,
) -> None:
    """Pós-Fase 3: as 11 colunas após a 9ª (advogados_consultados_escritorio,
    observacoes, id, hash, numero_processo, idOrgao, codigoClasse,
    numeroComunicacao, destinatarios, destinatarioadvogados,
    oabs_externas_consultadas) ficam exportadas mas com
    ``column_dimensions[letter].hidden=True`` no xlsx — Excel as esconde
    por default, usuário pode mostrar via menu se precisar."""
    from openpyxl.utils import get_column_letter

    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx
    from notion_rpadv.services.dje_transform import (
        CANONICAL_COLUMNS,
        HIDDEN_COLUMNS,
    )

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    # HIDDEN_COLUMNS é o contrato; cada uma deve estar marcada hidden=True.
    assert len(HIDDEN_COLUMNS) == 11
    for col_idx, name in enumerate(CANONICAL_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if name in HIDDEN_COLUMNS:
            assert dim is not None and dim.hidden, (
                f"Coluna {col_idx} ({name!r}) deveria estar oculta"
            )
        else:
            # Visíveis: ou não tem dimensão (default visível) ou
            # explicitamente hidden=False.
            if dim is not None:
                assert not dim.hidden, (
                    f"Coluna {col_idx} ({name!r}) deveria estar visível"
                )


def test_pos_F3_data_disponibilizacao_underscore_nao_aparece_no_xlsx(
    tmp_path: Path,
) -> None:
    """Pós-Fase 3: a coluna ``data_disponibilizacao`` (com underscore)
    SAIU do xlsx — só ``datadisponibilizacao`` (sem underscore, formato
    BR DD/MM/AAAA) aparece. SQLite ainda guarda a versão ISO internamente,
    e ``sort_rows`` ainda usa pra ordenação, mas o Excel só mostra a
    forma BR (que é o que o operador lê)."""
    from notion_rpadv.services.dje_exporter import write_publicacoes_xlsx

    rows = [{
        "id": 1, "hash": "h1", "siglaTribunal": "TRT10",
        "data_disponibilizacao": "2026-04-30",
        "datadisponibilizacao": "30/04/2026",
        "advogado_consultado": "Leonardo Guedes da Fonseca Passos (36129/DF)",
        "destinatarioadvogados": [],
    }]
    result = write_publicacoes_xlsx(
        rows, tmp_path, date(2026, 4, 30), date(2026, 4, 30),
    )
    wb = _read_back(result.path)
    ws = wb["Publicacoes"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
    ]
    assert "data_disponibilizacao" not in headers
    assert "datadisponibilizacao" in headers
    # E a célula da datadisponibilizacao tem o valor BR formatado.
    di_idx = headers.index("datadisponibilizacao") + 1
    assert ws.cell(row=2, column=di_idx).value == "30/04/2026"
