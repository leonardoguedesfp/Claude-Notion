"""Round 4 Frente 4 — testes pro snapshot exporter (xlsx).

Cobre helpers internos (_extract_title, _build_title_cache,
_format_for_excel) + integração end-to-end com NotionClient mockado.
Não testa a UI (ExportarPage) aqui — testes de Qt headless ficam
no test_round_4_layout.py."""
from __future__ import annotations

import datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest


# ---------------------------------------------------------------------------
# Helpers internos: _extract_title, _build_title_cache, _format_for_excel
# ---------------------------------------------------------------------------


def _page(page_id: str, title: str = "", **other_props) -> dict:
    """Mock de page response da API Notion. ``title`` vira a propriedade
    'Nome' (type=title); ``other_props`` adiciona props extras com type
    inferido."""
    properties: dict[str, dict] = {
        "Nome": {
            "type": "title",
            "title": [{"plain_text": title}] if title else [],
        },
    }
    properties.update(other_props)
    return {"id": page_id, "properties": properties}


def test_R4_F4_extract_title_returns_concatenated_plain_text() -> None:
    from notion_rpadv.services.snapshot_exporter import _extract_title
    page = {
        "id": "p1",
        "properties": {
            "Nome": {
                "type": "title",
                "title": [
                    {"plain_text": "João "},
                    {"plain_text": "da Silva"},
                ],
            },
        },
    }
    assert _extract_title(page) == "João da Silva"


def test_R4_F4_extract_title_empty_when_no_title_property() -> None:
    from notion_rpadv.services.snapshot_exporter import _extract_title
    page = {"id": "p1", "properties": {"Status": {"type": "select"}}}
    assert _extract_title(page) == ""


def test_R4_F4_build_title_cache_collects_all_bases() -> None:
    from notion_rpadv.services.snapshot_exporter import _build_title_cache
    cache = _build_title_cache({
        "Clientes": [_page("c1", "Maria"), _page("c2", "João")],
        "Processos": [_page("p1", "0001-00")],
    })
    assert cache == {"c1": "Maria", "c2": "João", "p1": "0001-00"}


def test_R4_F4_format_for_excel_none() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(None, "select", {}, {})
    assert val is None and miss == 0


def test_R4_F4_format_for_excel_checkbox_true_returns_sim() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(True, "checkbox", {}, {})
    assert val == "Sim" and miss == 0


def test_R4_F4_format_for_excel_checkbox_false_returns_none() -> None:
    """False → None pra deixar célula vazia (regra do spec)."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(False, "checkbox", {}, {})
    assert val is None and miss == 0


def test_R4_F4_format_for_excel_date_iso_returns_date_object() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel("2025-03-20", "date", {}, {})
    assert val == datetime.date(2025, 3, 20)
    assert miss == 0


def test_R4_F4_format_for_excel_date_with_time_extracts_date_part() -> None:
    """Notion pode retornar datetime com timezone (ex.: '2025-03-20T14:30:00-03:00')."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel("2025-03-20T14:30:00-03:00", "date", {}, {})
    assert val == datetime.date(2025, 3, 20)
    assert miss == 0


def test_R4_F4_format_for_excel_date_invalid_falls_back_to_string() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel("not-a-date", "date", {}, {})
    assert val == "not-a-date"  # fallback legível
    assert miss == 0


def test_R4_F4_format_for_excel_multi_select_joins_with_comma() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(
        ["Horas extras", "Indenização — I"], "multi_select", {}, {},
    )
    assert val == "Horas extras, Indenização — I"
    assert miss == 0


def test_R4_F4_format_for_excel_people_resolves_known_uuid_to_name() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    notion_users = {"uuid-1": {"name": "Déborah", "initials": "DM"}}
    val, miss = _format_for_excel(["uuid-1"], "people", {}, notion_users)
    assert val == "Déborah"
    assert miss == 0


def test_R4_F4_format_for_excel_people_falls_back_to_uuid_for_unknown() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(
        ["uuid-desconhecido"], "people", {}, {},
    )
    assert val == "uuid-desconhecido"
    assert miss == 0


def test_R4_F4_format_for_excel_relation_resolves_known_uuid_to_title() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    title_cache = {"uuid-1": "Maria Silva"}
    val, miss = _format_for_excel(["uuid-1"], "relation", title_cache, {})
    assert val == "Maria Silva"
    assert miss == 0


def test_R4_F4_format_for_excel_relation_marks_missing_with_question() -> None:
    """UUID fora do snapshot vira '[?]' e conta como miss pra warning."""
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(["uuid-fora"], "relation", {}, {})
    assert val == "[?]"
    assert miss == 1


def test_R4_F4_format_for_excel_relation_multiple_joins_with_comma() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    title_cache = {"u1": "A", "u2": "B"}
    val, miss = _format_for_excel(["u1", "u2"], "relation", title_cache, {})
    assert val == "A, B"
    assert miss == 0


def test_R4_F4_format_for_excel_number_passes_through() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel(42.5, "number", {}, {})
    assert val == 42.5 and miss == 0


def test_R4_F4_format_for_excel_text_passes_through() -> None:
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    val, miss = _format_for_excel("hello", "rich_text", {}, {})
    assert val == "hello" and miss == 0


# ---------------------------------------------------------------------------
# Integração end-to-end: export_snapshot com mock client + schema injetado
# ---------------------------------------------------------------------------


def _fake_schema_registry(schemas: dict[str, dict]) -> object:
    """Retorna um stub que satisfaz o uso de SchemaRegistry pelo exporter
    (acessa _schemas via private attr; o stub expõe apenas isso)."""
    stub = MagicMock()
    stub._schemas = schemas
    return stub


def _make_schema(properties: list[tuple[str, str]]) -> dict:
    """Helper: schema canônico no formato que o parser produziria.
    properties = [(notion_name, tipo), ...]"""
    props_dict = {}
    for order, (name, tipo) in enumerate(properties, start=1):
        slug = name.lower().replace(" ", "_").replace("-", "_")
        # NFKD seria mais correto, mas pra teste basta
        props_dict[slug] = {
            "notion_name": name,
            "tipo": tipo,
            "label": name,
            "default_order": order,
            "default_visible": True,
            "opcoes": [],
        }
    return {"properties": props_dict}


def test_R4_F4_export_snapshot_writes_aux_sheet_first(tmp_path: Path) -> None:
    """Aba 'Como ler este arquivo' é a primeira do workbook (índice 0)."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    schemas = {"Clientes": _make_schema([("Nome", "title")])}
    reg = _fake_schema_registry(schemas)

    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Clientes"], dest_path=dest,
        schema_registry=reg, data_sources={"Clientes": "ds-c"},
        notion_users={},
    )
    wb = load_workbook(dest)
    sheet_names = wb.sheetnames
    assert sheet_names[0] == "Como ler este arquivo"
    assert "Clientes" in sheet_names


def test_R4_F4_export_snapshot_one_sheet_per_selected_base(
    tmp_path: Path,
) -> None:
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    schemas = {
        "Clientes":  _make_schema([("Nome", "title")]),
        "Processos": _make_schema([("Nome", "title")]),
        "Tarefas":   _make_schema([("Nome", "title")]),
    }
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Clientes", "Tarefas"],
        dest_path=dest,
        schema_registry=reg,
        data_sources={
            "Clientes": "ds-c", "Processos": "ds-p", "Tarefas": "ds-t",
        },
        notion_users={},
    )
    wb = load_workbook(dest)
    # Aux + 2 selecionadas — Processos NÃO deve estar.
    assert wb.sheetnames == ["Como ler este arquivo", "Clientes", "Tarefas"]


def test_R4_F4_export_snapshot_header_row_uses_notion_names(
    tmp_path: Path,
) -> None:
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    # Round 5: usa base sintética não coberta pelo DEFAULT_LAYOUTS pra que
    # a ordem das colunas seja a do schema (caso fallback). O teste
    # ordering-aware via layout fica em test_round_5_rendering.
    schemas = {"Synth": _make_schema([
        ("Nome",    "title"),
        ("E-mail",  "email"),
        ("Telefone", "phone_number"),
    ])}
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Synth"], dest_path=dest,
        schema_registry=reg, data_sources={"Synth": "ds-s"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Synth"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
    assert headers == ["Nome", "E-mail", "Telefone"]
    # Headers em negrito
    assert ws.cell(row=1, column=1).font.bold is True


def test_R4_F4_export_snapshot_writes_data_rows_with_typed_values(
    tmp_path: Path,
) -> None:
    """Dados são escritos com valores tipados: number como número, date
    como date, checkbox como Sim/None, etc."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    pages = [
        {
            "id": "c1",
            "properties": {
                "Nome": {
                    "type": "title",
                    "title": [{"plain_text": "Maria"}],
                },
                "Adiantamento": {"type": "checkbox", "checkbox": True},
                "Data de aposentadoria": {
                    "type": "date",
                    "date": {"start": "2030-01-15"},
                },
                "Idade": {"type": "number", "number": 65},
                "Areas": {
                    "type": "multi_select",
                    "multi_select": [
                        {"name": "Cível"}, {"name": "Trabalhista"},
                    ],
                },
            },
        },
    ]
    client = MagicMock()
    client.query_all.return_value = pages
    # Round 5: base sintética fora de DEFAULT_LAYOUTS preserva ordem do
    # schema (relevante pras assertivas posicionais abaixo).
    schemas = {"Synth": _make_schema([
        ("Nome",                  "title"),
        ("Adiantamento",          "checkbox"),
        ("Data de aposentadoria", "date"),
        ("Idade",                 "number"),
        ("Areas",                 "multi_select"),
    ])}
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Synth"], dest_path=dest,
        schema_registry=reg, data_sources={"Synth": "ds-s"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["Synth"]
    assert ws.cell(row=2, column=1).value == "Maria"
    assert ws.cell(row=2, column=2).value == "Sim"  # checkbox True
    # openpyxl coerce datetime.date pra datetime.datetime na leitura;
    # comparamos a parte de data pra acomodar ambos.
    date_val = ws.cell(row=2, column=3).value
    if isinstance(date_val, datetime.datetime):
        date_val = date_val.date()
    assert date_val == datetime.date(2030, 1, 15)
    assert ws.cell(row=2, column=4).value == 65
    assert ws.cell(row=2, column=5).value == "Cível, Trabalhista"


def test_R4_F4_export_snapshot_resolves_relation_titles_in_xlsx(
    tmp_path: Path,
) -> None:
    """Relations apontando pra page_ids dentro do snapshot são resolvidos
    pelos títulos correspondentes no xlsx (não UUIDs)."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    cliente_page = {
        "id": "cli-1",
        "properties": {
            "Nome": {"type": "title", "title": [{"plain_text": "João"}]},
        },
    }
    tarefa_page = {
        "id": "tar-1",
        "properties": {
            "Nome": {"type": "title", "title": [{"plain_text": "T1"}]},
            "Cliente": {
                "type": "relation",
                "relation": [{"id": "cli-1"}],
            },
        },
    }
    # Mock client retorna pages diferentes por dsid
    def query_all(dsid: str, on_progress=None):
        if dsid == "ds-c":
            return [cliente_page]
        if dsid == "ds-t":
            return [tarefa_page]
        return []
    client = MagicMock()
    client.query_all.side_effect = query_all
    # Round 5: bases sintéticas fora do DEFAULT_LAYOUTS preservam ordem
    # do schema. SynthClientes pra resolver relation; SynthTarefas tem
    # (Nome, Cliente) na ordem do schema.
    schemas = {
        "SynthClientes": _make_schema([("Nome", "title")]),
        "SynthTarefas":  _make_schema([
            ("Nome", "title"), ("Cliente", "relation"),
        ]),
    }
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    result = export_snapshot(
        client=client, bases=["SynthClientes", "SynthTarefas"],
        dest_path=dest, schema_registry=reg,
        data_sources={"SynthClientes": "ds-c", "SynthTarefas": "ds-t"},
        notion_users={},
    )
    wb = load_workbook(dest)
    ws = wb["SynthTarefas"]
    # col 2 = Cliente (relation). Resolvido pra título de cli-1.
    assert ws.cell(row=2, column=2).value == "João"
    assert result.relation_misses == 0


def test_R4_F4_export_snapshot_marks_missing_relation_uuid(
    tmp_path: Path,
) -> None:
    """Relations apontando pra page_ids fora do snapshot viram '[?]'
    e contam no relation_misses."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot

    pages = [{
        "id": "tar-1",
        "properties": {
            "Nome": {"type": "title", "title": [{"plain_text": "T1"}]},
            "Cliente": {
                "type": "relation",
                "relation": [{"id": "cli-uuid-fora"}],
            },
        },
    }]
    client = MagicMock()
    client.query_all.return_value = pages
    schemas = {"Tarefas": _make_schema([("Nome", "title"), ("Cliente", "relation")])}
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    result = export_snapshot(
        client=client, bases=["Tarefas"], dest_path=dest,
        schema_registry=reg, data_sources={"Tarefas": "ds-t"},
        notion_users={},
    )
    assert result.relation_misses == 1


def test_R4_F4_export_snapshot_aux_sheet_has_counts_per_base(
    tmp_path: Path,
) -> None:
    """Aba auxiliar lista contagem por base. Não asseguro layout exato
    (cell coordinates) — só a presença das contagens."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    pages_3 = [_page(f"p{i}", f"P{i}") for i in range(3)]
    pages_5 = [_page(f"q{i}", f"Q{i}") for i in range(5)]

    def query_all(dsid: str, on_progress=None):
        return pages_3 if dsid == "ds-c" else pages_5
    client = MagicMock()
    client.query_all.side_effect = query_all
    schemas = {
        "Clientes": _make_schema([("Nome", "title")]),
        "Tarefas":  _make_schema([("Nome", "title")]),
    }
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    result = export_snapshot(
        client=client, bases=["Clientes", "Tarefas"],
        dest_path=dest, schema_registry=reg,
        data_sources={"Clientes": "ds-c", "Tarefas": "ds-t"},
        notion_users={},
    )
    assert result.counts == {"Clientes": 3, "Tarefas": 5}
    wb = load_workbook(dest)
    aux = wb["Como ler este arquivo"]
    # Procura células com 3 e 5 e os labels — tolerante ao layout.
    cells = [c.value for row in aux.iter_rows() for c in row]
    assert "Clientes" in cells
    assert "Tarefas" in cells
    assert 3 in cells
    assert 5 in cells


def test_R4_F4_export_snapshot_calls_progress_during_fetch_and_write(
    tmp_path: Path,
) -> None:
    """on_progress é chamado em ambas as fases (fetch e write)."""
    from notion_rpadv.services.snapshot_exporter import (
        PHASE_FETCH,
        PHASE_WRITE,
        export_snapshot,
    )

    pages = [_page(f"p{i}", f"P{i}") for i in range(3)]
    client = MagicMock()

    # query_all simula chamadas de progresso durante o fetch
    def query_all(dsid: str, on_progress=None):
        if on_progress is not None:
            on_progress(1)
            on_progress(3)
        return pages
    client.query_all.side_effect = query_all
    schemas = {"Clientes": _make_schema([("Nome", "title")])}
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")

    progress_log: list[tuple[str, str, int, int]] = []
    export_snapshot(
        client=client, bases=["Clientes"], dest_path=dest,
        on_progress=lambda b, ph, c, t: progress_log.append((b, ph, c, t)),
        schema_registry=reg, data_sources={"Clientes": "ds-c"},
        notion_users={},
    )
    phases = {p[1] for p in progress_log}
    assert PHASE_FETCH in phases
    assert PHASE_WRITE in phases
    # Pelo menos uma entrada de cada fase pra Clientes
    assert any(p[0] == "Clientes" and p[1] == PHASE_FETCH for p in progress_log)
    assert any(p[0] == "Clientes" and p[1] == PHASE_WRITE for p in progress_log)


def test_R4_F4_export_snapshot_silently_skips_unknown_base(
    tmp_path: Path,
) -> None:
    """Base solicitada mas ausente em data_sources é silenciosamente
    ignorada (sem crash)."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    from openpyxl import load_workbook

    client = MagicMock()
    client.query_all.return_value = []
    schemas = {"Clientes": _make_schema([("Nome", "title")])}
    reg = _fake_schema_registry(schemas)
    dest = str(tmp_path / "out.xlsx")
    export_snapshot(
        client=client, bases=["Clientes", "Inexistente"], dest_path=dest,
        schema_registry=reg, data_sources={"Clientes": "ds-c"},
        notion_users={},
    )
    wb = load_workbook(dest)
    assert "Inexistente" not in wb.sheetnames


def test_R4_F4_export_snapshot_requires_token_or_client() -> None:
    """Sem token nem client, levanta ValueError."""
    from notion_rpadv.services.snapshot_exporter import export_snapshot
    with pytest.raises(ValueError):
        export_snapshot(
            bases=["Clientes"], dest_path="/tmp/x.xlsx",
        )


def test_R4_F4_exportar_page_imports_and_has_4_checkboxes() -> None:
    """Smoke: ExportarPage importa e tem checkboxes pra cada base
    em DATA_SOURCES."""
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)

    from notion_bulk_edit.config import DATA_SOURCES
    from notion_rpadv.pages.exportar import ExportarPage
    page = ExportarPage(
        conn=MagicMock(), token="dummy", user="leo",
    )
    # Cada base de DATA_SOURCES tem checkbox
    for base in DATA_SOURCES:
        assert base in page._checkboxes  # noqa: SLF001
        assert page._checkboxes[base].isChecked() is True  # noqa: SLF001


# ---------------------------------------------------------------------------
# Hotfix pós-merge — defaults respeitados + guard com toast amigável
# ---------------------------------------------------------------------------


def _make_exportar_page():
    """Helper: instancia ExportarPage com QApplication garantida."""
    import sys
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication(sys.argv)
    from notion_rpadv.pages.exportar import ExportarPage
    return ExportarPage(conn=MagicMock(), token="dummy", user="leo")


def test_R4_F4_exportar_page_starts_with_all_bases_checked() -> None:
    """Round 4 hotfix: estado inicial dos checkboxes é Qt.CheckState.Checked
    (não só isChecked() — defende contra tristate / partial-checked que
    confundiria handlers que usam checkState() == Checked)."""
    from PySide6.QtCore import Qt
    from notion_bulk_edit.config import DATA_SOURCES
    page = _make_exportar_page()
    for base in DATA_SOURCES:
        cb = page._checkboxes[base]  # noqa: SLF001
        assert cb.isChecked() is True, f"{base} não está marcado"
        assert cb.checkState() == Qt.CheckState.Checked, (
            f"{base} em estado tristate/partial em vez de Checked"
        )


def test_R4_F4_exportar_page_collect_bases_returns_all_when_default() -> None:
    """Round 4 hotfix: ``_collect_bases`` retorna as 4 bases (em ordem do
    spec) sem que o operador toque em nada. Garantia funcional do
    "todas marcadas por default"."""
    from notion_bulk_edit.config import DATA_SOURCES
    page = _make_exportar_page()
    bases = page._collect_bases()  # noqa: SLF001
    # Todas presentes
    assert set(bases) == set(DATA_SOURCES.keys())
    # Ordem do spec — preservada via _BASE_ORDER
    assert bases == ["Clientes", "Processos", "Tarefas", "Catalogo"]


def test_R4_F4_exportar_page_empty_selection_shows_friendly_error() -> None:
    """Round 4 hotfix: nenhum base marcado + clique em Exportar →
    toast de warning amigável + nenhum worker disparado. Sem isso, um
    bug que esvaziasse a lista chegaria no exporter como ValueError
    genérico do openpyxl."""
    page = _make_exportar_page()
    # Desmarca todos
    for cb in page._checkboxes.values():  # noqa: SLF001
        cb.setChecked(False)
    # Captura toasts emitidos
    toasts: list[tuple[str, str]] = []
    page.toast_requested.connect(
        lambda msg, kind: toasts.append((msg, kind)),
    )
    # Confirma estado pré: bases vazio
    assert page._collect_bases() == []  # noqa: SLF001
    # Click handler (curto-circuita antes do QFileDialog)
    page._on_export_clicked()  # noqa: SLF001
    # Worker NÃO iniciado
    assert page._thread is None  # noqa: SLF001
    assert page._worker is None  # noqa: SLF001
    # Toast emitido com kind=warning e mensagem reconhecível
    assert len(toasts) == 1
    msg, kind = toasts[0]
    assert kind == "warning"
    assert "selecione" in msg.lower()


def test_R4_F4_format_for_excel_handles_list_in_catchall() -> None:
    """Round 4 hotfix: rollup com type="array" decodifica como list. O
    catch-all do _format_for_excel agora joina lists em string em vez de
    devolver list pro openpyxl (que raise 'Cannot convert [] to Excel').

    Isso é o fix real do bug ValueError observado pelo operador no
    smoke pós-merge.
    """
    from notion_rpadv.services.snapshot_exporter import _format_for_excel
    # Lista vazia (rollup array sem itens) → string vazia, não [].
    val, miss = _format_for_excel([], "rollup", {}, {})
    assert val == ""
    assert miss == 0
    # Lista com itens → comma-separated string.
    val, miss = _format_for_excel([1, 2, 3], "rollup", {}, {})
    assert val == "1, 2, 3"
    assert miss == 0
    # Lista com None misturado → None é descartado.
    val, miss = _format_for_excel(["a", None, "b"], "rollup", {}, {})
    assert val == "a, b"
    assert miss == 0
