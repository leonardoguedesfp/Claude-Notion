"""Testes de integração end-to-end pós-Fase 3 hotfix watermark integrity.

Refator (2026-05-02): cursor único → cursor por advogado em
``djen_advogado_state``. Worker recebe ``list[AdvogadoConsulta]`` em
vez de ``advogados, data_inicio, data_fim`` global. Cada advogado tem
janela individual e cursor próprio que avança independentemente.

Cenários:
- Sucesso completo (todos os 6 avançam pra data_fim de suas janelas)
- Falha parcial (advogados que falharam mantêm cursor; outros avançam)
- Modo manual não toca em ``djen_advogado_state``
- Migração legada
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

from notion_rpadv.services import dje_db, dje_state
from notion_rpadv.services.dje_advogados import ADVOGADOS
from notion_rpadv.services.dje_client import AdvogadoConsulta, AdvogadoResult


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _raw_row(rid: int, *, data="2026-04-30", oab="36129", uf="DF") -> dict:
    nome = "Leonardo Guedes da Fonseca Passos" if oab == "36129" else (
        "Ricardo Luiz Rodrigues da Fonseca Passos" if oab == "15523"
        else "Externo"
    )
    return {
        "id": rid,
        "hash": f"hash{rid:04d}",
        "siglaTribunal": "TRT10",
        "data_disponibilizacao": data,
        "datadisponibilizacao": "30/04/2026",
        "numero_processo": f"00012345-6.2026.{rid}",
        "tipoComunicacao": "Intimação",
        "destinatarios": [],
        "destinatarioadvogados": [
            {"advogado": {"nome": nome, "numero_oab": oab, "uf_oab": uf}},
        ],
        "ativo": True,
        "status": "P",
        "meio": "D",
        "meiocompleto": "Diário de Justiça Eletrônico Nacional",
        "advogado_consultado": f"{nome} ({oab}/{uf})",
    }


def _make_summary(
    rows: list[dict],
    *,
    consultas: list[AdvogadoConsulta],
    failed: list[dict] | None = None,
    cancelled: bool = False,
):
    """FetchSummary mockado com ``by_advogado`` populado por
    ``AdvogadoResult`` reais. ``failed`` é a lista de advogados (dict)
    que devem aparecer com erro + ``data_max_safe`` antes da janela."""
    failed_oabs = {(a["oab"], a["uf"]) for a in (failed or [])}
    by_adv = []
    for c in consultas:
        ar = AdvogadoResult(advogado=c.advogado, sub_windows=[(c.data_inicio, c.data_fim)])
        if (c.advogado["oab"], c.advogado["uf"]) in failed_oabs:
            ar.erro = "fake fail"
            ar.data_max_safe = date(2025, 12, 31)
        else:
            ar.data_max_safe = c.data_fim
        by_adv.append(ar)
    summary = MagicMock()
    summary.rows = rows
    summary.cancelled = cancelled
    summary.total_items = len(rows)
    summary.by_advogado = by_adv
    summary.errors = [r for r in by_adv if r.erro is not None]
    return summary


def _run_worker(
    *,
    conn,
    rows,
    output_dir: Path,
    consultas: list[AdvogadoConsulta] | None = None,
    failed: list[dict] | None = None,
    cancelled: bool = False,
    mode: str = "padrao",
    oabs_externas: set[str] | None = None,
):
    """Constrói _DJEWorker com FetchSummary mockado e roda síncrono."""
    from notion_rpadv.pages.leitor_dje import _DJEWorker

    if consultas is None:
        consultas = [
            AdvogadoConsulta(
                advogado=a,
                data_inicio=date(2026, 4, 1),
                data_fim=date(2026, 4, 30),
            )
            for a in ADVOGADOS
        ]
    marcadas = (
        {f"{c.advogado['oab']}/{c.advogado['uf']}" for c in consultas}
        if mode == "padrao"
        else set()
    )
    if oabs_externas is None:
        oabs_externas = set()
    worker = _DJEWorker(
        consultas=consultas,
        output_dir=output_dir,
        mode=mode,
        oabs_escritorio_marcadas=marcadas,
        oabs_externas_pesquisadas=oabs_externas,
        dje_conn=conn,
    )
    summary = _make_summary(
        rows, consultas=consultas, failed=failed, cancelled=cancelled,
    )
    outcomes: list = []
    worker.finished.connect(lambda o: outcomes.append(o))
    with patch(
        "notion_rpadv.pages.leitor_dje.DJEClient",
    ) as mock_client_cls:
        mock_client = MagicMock()
        mock_client.fetch_all.return_value = summary
        mock_client_cls.return_value = mock_client
        worker._run_inner()  # noqa: SLF001
    assert outcomes, "worker.finished não emitiu"
    return outcomes[0]


# ---------------------------------------------------------------------------
# Cenário básico: cursor por advogado avança independente
# ---------------------------------------------------------------------------


def test_modo_padrao_avanca_cursor_de_cada_advogado(tmp_path: Path) -> None:
    """Modo padrão com sucesso completo: cursor de CADA advogado avança
    pra data_fim da sua janela individual."""
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        rows = [_raw_row(rid=i) for i in range(1, 6)]
        out = _run_worker(
            conn=conn, rows=rows, output_dir=output_dir,
        )
        # Todos os 6 advogados têm cursor = 30/04/2026
        for adv in ADVOGADOS:
            assert dje_state.read_advogado_cursor(
                conn, oab=adv["oab"], uf=adv["uf"],
            ) == date(2026, 4, 30)
        assert out.novas_inseridas == 5
    finally:
        conn.close()


def test_falha_parcial_segura_apenas_o_advogado_falhado(tmp_path: Path) -> None:
    """5 advogados sucesso + 1 falha → 5 cursores avançam pra data_fim,
    1 cursor (do que falhou) NÃO atualiza pq data_max_safe = data_inicio - 1d
    (regressão garantida)."""
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        # Pré-popula cursor de Ricardo (vai ser o que falha)
        dje_state.update_advogado_cursor(
            conn, oab="15523", uf="DF", novo_cursor=date(2026, 3, 31),
        )
        ricardo = next(a for a in ADVOGADOS if a["oab"] == "15523")
        rows = [_raw_row(rid=1)]
        out = _run_worker(
            conn=conn, rows=rows, output_dir=output_dir,
            failed=[ricardo],
        )
        # Ricardo: cursor mantido (anti-regressão)
        assert dje_state.read_advogado_cursor(
            conn, oab="15523", uf="DF",
        ) == date(2026, 3, 31)
        # Outros 5: cursor avançou pra 30/04
        for adv in ADVOGADOS:
            if adv["oab"] == "15523":
                continue
            assert dje_state.read_advogado_cursor(
                conn, oab=adv["oab"], uf=adv["uf"],
            ) == date(2026, 4, 30)
        # Banner reporta 1 erro
        assert len(out.errors) == 1
        assert out.errors[0]["oab"] == "15523"
    finally:
        conn.close()


def test_falha_principal_segura_cursor_mesmo_se_retry_recupera(
    tmp_path: Path,
) -> None:
    """Pós-smoke real do refator (2026-05-02): valida que retry diferido
    NÃO destranca o cursor. Cenário: advogado falhou na varredura
    PRINCIPAL → cursor mantém valor anterior, mesmo se retry diferido
    eventualmente recuperar items.

    Reproduz literalmente o cenário do smoke real onde 5 advogados
    falharam com 429 mas todos avançaram pra ``data_fim``. A correção
    é: ``data_max_safe`` é calculado a partir de
    ``failed_per_adv_main`` (snapshot pré-retry), não do estado pós-
    retry. Logo, falha na principal trava o cursor independente do
    retry recuperar ou não.
    """
    from notion_rpadv.services.dje_client import (
        AdvogadoConsulta, AdvogadoResult,
    )

    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        # Pré-popula cursor pra Ricardo
        dje_state.update_advogado_cursor(
            conn, oab="15523", uf="DF", novo_cursor=date(2026, 3, 1),
        )

        # Cenário: Ricardo falhou na varredura principal (sub_idx=0).
        # Mesmo se retry recuperasse, ``data_max_safe`` deve ser
        # ``data_inicio - 1d`` (snapshot pré-retry capturou a falha).
        # Aqui simulamos via summary mockado.
        ricardo = next(a for a in ADVOGADOS if a["oab"] == "15523")

        ar = AdvogadoResult(advogado=ricardo)
        ar.erro = "HTTP 429"  # erro permanece (não é limpo pelo retry)
        ar.data_max_safe = date(2025, 12, 31)  # data_inicio - 1d
        # Items podem ter sido recuperados pelo retry — vão pro banco
        # mas não destrancam o cursor.

        summary = MagicMock()
        summary.rows = []  # nenhuma publicação no payload
        summary.cancelled = False
        summary.total_items = 0
        summary.by_advogado = [ar]
        summary.errors = [ar]

        from notion_rpadv.pages.leitor_dje import _DJEWorker

        worker = _DJEWorker(
            consultas=[AdvogadoConsulta(
                advogado=ricardo,
                data_inicio=date(2026, 1, 1),
                data_fim=date(2026, 4, 30),
            )],
            output_dir=output_dir,
            mode="padrao",
            oabs_escritorio_marcadas={"15523/DF"},
            oabs_externas_pesquisadas=set(),
            dje_conn=conn,
        )
        outcomes: list = []
        worker.finished.connect(lambda o: outcomes.append(o))
        with patch(
            "notion_rpadv.pages.leitor_dje.DJEClient",
        ) as mock_client_cls:
            mock_client = MagicMock()
            mock_client.fetch_all.return_value = summary
            mock_client_cls.return_value = mock_client
            worker._run_inner()  # noqa: SLF001
        out = outcomes[0]

        # Cursor de Ricardo NÃO avançou pra data_fim — preservou
        # o valor anterior (anti-regressão do update_advogado_cursor).
        assert dje_state.read_advogado_cursor(
            conn, oab="15523", uf="DF",
        ) == date(2026, 3, 1)
        # Banner reporta o erro (semântica conservadora).
        assert len(out.errors) == 1
        assert out.errors[0]["oab"] == "15523"
    finally:
        conn.close()


def test_modo_manual_nao_toca_banco_nem_state(tmp_path: Path) -> None:
    """Modo manual (refator pós-Fase 3 hotfix UX) é COMPLETAMENTE
    transient: NÃO escreve em ``publicacoes``, NÃO toca
    ``djen_advogado_state``, NÃO regenera ``Historico_DJEN_completo.xlsx``.
    Só gera o Excel-de-execução.
    """
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        dje_state.update_advogado_cursor(
            conn, oab="15523", uf="DF", novo_cursor=date(2026, 3, 1),
        )
        # Pré-popula 1 publicação no banco pra ter referência inicial
        dje_db.insert_publicacao(
            conn, djen_id=42, hash_="hpre",
            oabs_escritorio="Pre (15523/DF)", oabs_externas="",
            numero_processo=None, data_disponibilizacao="2026-03-01",
            sigla_tribunal="TRT10", payload={"id": 42}, mode="padrao",
        )
        conn.commit()
        count_inicial = dje_db.count_publicacoes(conn)

        ext_consulta = AdvogadoConsulta(
            advogado={"nome": "", "oab": "99999", "uf": "SP"},
            data_inicio=date(2026, 2, 1),
            data_fim=date(2026, 2, 28),
        )
        ext_row = _raw_row(rid=99, data="2026-02-15", oab="99999", uf="SP")
        out = _run_worker(
            conn=conn, rows=[ext_row], output_dir=output_dir,
            consultas=[ext_consulta], mode="manual",
            oabs_externas={"99999/SP"},
        )
        # Cursor de Ricardo NÃO mudou
        assert dje_state.read_advogado_cursor(
            conn, oab="15523", uf="DF",
        ) == date(2026, 3, 1)
        # Banco intacto — manual NÃO insere
        assert dje_db.count_publicacoes(conn) == count_inicial
        # State não tem entrada pra OAB externa
        all_state = dje_state.read_all_advogados_state(conn)
        assert ("99999", "SP") not in all_state
        # Excel-de-execução foi gerado mesmo assim
        assert out.excel_path is not None
        assert out.excel_path.exists()
        # Histórico NÃO foi regenerado em modo manual
        assert out.historico_path is None
        # count_antes = 0 (N/A no manual); novas_inseridas = publicações captadas
        assert out.count_antes == 0
        assert out.novas_inseridas == 1
        assert out.mode == "manual"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Janela individual por advogado
# ---------------------------------------------------------------------------


def test_consultas_com_janelas_diferentes(tmp_path: Path) -> None:
    """Cada advogado pode ter sua própria janela. Após a varredura,
    cursor de cada um vai pra data_fim da sua janela."""
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        # 2 consultas com janelas diferentes
        adv_a = ADVOGADOS[0]
        adv_b = ADVOGADOS[1]
        consultas = [
            AdvogadoConsulta(
                advogado=adv_a,
                data_inicio=date(2026, 4, 1),
                data_fim=date(2026, 4, 30),
            ),
            AdvogadoConsulta(
                advogado=adv_b,
                data_inicio=date(2026, 1, 1),
                data_fim=date(2026, 4, 30),
            ),
        ]
        _run_worker(
            conn=conn, rows=[], output_dir=output_dir,
            consultas=consultas,
        )
        assert dje_state.read_advogado_cursor(
            conn, oab=adv_a["oab"], uf=adv_a["uf"],
        ) == date(2026, 4, 30)
        assert dje_state.read_advogado_cursor(
            conn, oab=adv_b["oab"], uf=adv_b["uf"],
        ) == date(2026, 4, 30)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Excel-de-execução SEMPRE gerado (refator pós-Fase 3)
# ---------------------------------------------------------------------------


def test_excel_de_execucao_gerado_mesmo_com_zero_novas(tmp_path: Path) -> None:
    """Refator: Excel sempre gerado pra evidência. Inclui abas Status + Log."""
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        out = _run_worker(
            conn=conn, rows=[], output_dir=output_dir,
        )
        assert out.novas_inseridas == 0
        assert out.excel_path is not None
        assert out.excel_path.exists()
        # Abas presentes
        from openpyxl import load_workbook
        wb = load_workbook(out.excel_path)
        assert "Publicacoes" in wb.sheetnames
        assert "Status" in wb.sheetnames
        assert "Log" in wb.sheetnames
    finally:
        conn.close()


def test_aba_status_lista_todos_advogados_oficiais(tmp_path: Path) -> None:
    """Aba Status tem 1 linha por advogado oficial, com cursor + dias atrás."""
    db_path = tmp_path / "leitor_dje.db"
    output_dir = tmp_path / "saidas"
    conn = dje_db.get_connection(db_path)
    try:
        out = _run_worker(
            conn=conn, rows=[], output_dir=output_dir,
        )
        from openpyxl import load_workbook
        wb = load_workbook(out.excel_path)
        ws = wb["Status"]
        # Header + 6 linhas (1 por advogado oficial)
        assert ws.max_row == 1 + len(ADVOGADOS)
        # Headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers == [
            "advogado", "oab_uf", "ultimo_cursor",
            "dias_atras", "ultima_execucao",
        ]
        # Cada linha tem nome + oab/uf
        for row_idx, adv in enumerate(ADVOGADOS, start=2):
            assert ws.cell(row=row_idx, column=1).value == adv["nome"]
            assert ws.cell(row=row_idx, column=2).value == (
                f"{adv['oab']}/{adv['uf']}"
            )
    finally:
        conn.close()
