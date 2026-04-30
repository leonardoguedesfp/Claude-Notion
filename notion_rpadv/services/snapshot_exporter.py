"""Round 4 Frente 4 — snapshot exporter (xlsx) das bases do Notion.

Gera um arquivo ``.xlsx`` com:
- 1 aba auxiliar "Como ler este arquivo" (à esquerda) com metadados do
  snapshot (data/hora, contagens por base, legenda, aviso).
- 1 aba por base selecionada com cabeçalho em negrito (nomes de
  propriedades do Notion, na ordem da API) e uma linha por página da
  base com TODOS os campos (ignora visibilidade do app, ignora filtros).

Formatação por tipo:
- title / rich_text / url / email / phone_number → texto.
- number → number nativo.
- date → ``datetime.date`` (Excel exibe nativamente; sem formatação cromática).
- checkbox → "Sim" se marcado, vazio caso contrário.
- select / status → nome da opção como texto, sem fundo colorido.
- multi_select → valores separados por vírgula em uma única célula.
- relation → títulos das páginas relacionadas (resolução por UUID
  via cache built das bases selecionadas).
- people → nomes de NOTION_USERS quando reconhecido, UUIDs caso
  contrário (placeholders em config.NOTION_USERS para Mariana/Carla
  caem nesse fallback até OBS-A03 ser endereçada).
- created_time / last_edited_time → ISO string como texto (carrega o
  timezone do Notion; conversão para datetime exigiria parsing custom).

Pagination via ``NotionClient.query_all`` (limite 100/página, segue
``next_cursor``). Estimativa do spec: ~30-60s para as 4 bases (~2200
registros total).

API:
    export_snapshot(token, bases, dest_path, on_progress=None) -> ExportResult

Para testes, injete ``client`` e ``schema_registry``:
    export_snapshot(token=None, client=mock, schema_registry=reg, ...)
"""
from __future__ import annotations

import datetime as _dt
import time
from dataclasses import dataclass, field
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Phase markers usados em on_progress(base, phase, count, total).
PHASE_FETCH: str = "fetch"
PHASE_WRITE: str = "write"

_AUX_SHEET_NAME: str = "Como ler este arquivo"


@dataclass(frozen=True)
class ExportResult:
    """Retorno de export_snapshot — usado pela UI pra exibir resumo."""

    dest_path: str
    counts: dict[str, int] = field(default_factory=dict)
    duration_seconds: float = 0.0
    relation_misses: int = 0  # UUIDs em relations que não casaram no cache


def _extract_title(page: dict[str, Any]) -> str:
    """Retorna a string de título da página (concatena rich_text blocks
    do property type=title). String vazia se não encontra."""
    properties = page.get("properties", {})
    for _name, prop in properties.items():
        if prop.get("type") != "title":
            continue
        blocks = prop.get("title", []) or []
        return "".join(b.get("plain_text", "") for b in blocks)
    return ""


def _build_title_cache(
    pages_by_base: dict[str, list[dict[str, Any]]],
) -> dict[str, str]:
    """Coleta {page_id: title} pra cada page de cada base. Usado pra
    resolver relations dentro do snapshot sem chamadas adicionais à API
    (relations apontando pra páginas fora do snapshot caem em '[?]')."""
    cache: dict[str, str] = {}
    for _base, pages in pages_by_base.items():
        for page in pages:
            pid = page.get("id")
            if not pid:
                continue
            cache[str(pid)] = _extract_title(page)
    return cache


def _format_for_excel(
    value: Any,
    tipo: str,
    title_cache: dict[str, str],
    notion_users: dict[str, dict[str, str]],
) -> tuple[Any, int]:
    """Converte valor decoded do encoders em algo que openpyxl consegue
    escrever em célula. Retorna (cell_value, miss_count) onde miss_count
    é 1 quando uma relation referenciou UUID não cacheado.

    None → None (célula vazia). Listas viram strings comma-separated.
    Datas viram ``datetime.date``. Checkbox vira "Sim" ou None.
    """
    if value is None:
        return None, 0

    if tipo == "checkbox":
        return ("Sim" if value else None), 0

    if tipo == "date":
        # decode_value retorna ISO string (YYYY-MM-DD ou YYYY-MM-DDTHH:MM:SS...).
        # Tenta converter pro tipo date nativo do Excel; se falhar (datetime
        # com offset, etc.), cai pra string como fallback legível.
        if not isinstance(value, str):
            return value, 0
        try:
            # Date pura
            return _dt.date.fromisoformat(value), 0
        except ValueError:
            try:
                # Datetime com timezone — corta no T e tenta date
                return _dt.date.fromisoformat(value.split("T")[0]), 0
            except (ValueError, IndexError):
                return value, 0

    if tipo == "multi_select":
        if isinstance(value, list):
            return ", ".join(str(v) for v in value if v), 0
        return str(value), 0

    if tipo == "people":
        from notion_bulk_edit.config import resolve_user_name
        if not isinstance(value, list):
            return str(value), 0
        return ", ".join(
            resolve_user_name(u, users=notion_users) for u in value if u
        ), 0

    if tipo == "relation":
        if not isinstance(value, list):
            return str(value), 0
        misses = 0
        titles: list[str] = []
        for uuid in value:
            uuid_str = str(uuid)
            title = title_cache.get(uuid_str)
            if title is None or not title.strip():
                titles.append("[?]")
                misses += 1
            else:
                titles.append(title)
        return ", ".join(titles), misses

    # Catch-all: number, title/rich_text/url/email/phone, formula, rollup,
    # created/last_edited. Defesa: ``decode_value`` pra rollup com
    # type="array" retorna list (encoders.py:198-204). Sem este branch,
    # uma list vazia ([]) chegaria em ``ws.cell(value=[])`` e openpyxl
    # raise "Cannot convert [] to Excel". Round 4 hotfix pós-merge.
    if isinstance(value, list):
        return ", ".join(str(v) for v in value if v is not None), 0
    return value, 0


def _write_base_sheet(
    ws: Any,
    base: str,
    pages: list[dict[str, Any]],
    title_cache: dict[str, str],
    notion_users: dict[str, dict[str, str]],
    schema_registry: Any,
    on_progress: Callable[[str, str, int, int], None] | None,
) -> int:
    """Escreve uma aba com os dados de ``base``. Cabeçalho em negrito,
    1 linha por página, TODAS as propriedades do schema (ignora
    visibilidade do picker e filtros). Retorna miss_count de relations.
    """
    from notion_bulk_edit.encoders import decode_value

    parsed = schema_registry._schemas.get(base, {})  # noqa: SLF001
    properties = parsed.get("properties", {})
    headers: list[str] = []
    types: list[str] = []
    for _slug, prop_dict in properties.items():
        headers.append(prop_dict.get("notion_name") or _slug)
        types.append(prop_dict.get("tipo", "rich_text"))

    bold = Font(bold=True)
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
        cell.alignment = Alignment(horizontal="left")

    miss_total = 0
    total_pages = len(pages)
    for row_idx, page in enumerate(pages, start=2):
        props = page.get("properties", {}) or {}
        for col_idx, (header, tipo) in enumerate(zip(headers, types), start=1):
            prop_block = props.get(header, {})
            decoded = decode_value(prop_block, tipo)
            cell_value, miss = _format_for_excel(
                decoded, tipo, title_cache, notion_users,
            )
            miss_total += miss
            if cell_value is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if isinstance(cell_value, _dt.date) and not isinstance(
                    cell_value, _dt.datetime,
                ):
                    cell.number_format = "yyyy-mm-dd"
        if on_progress is not None and (row_idx - 2) % 50 == 0:
            on_progress(base, PHASE_WRITE, row_idx - 1, total_pages)
    if on_progress is not None:
        on_progress(base, PHASE_WRITE, total_pages, total_pages)
    return miss_total


def _write_aux_sheet(
    ws: Any,
    counts: dict[str, int],
    bases_in_order: list[str],
    snapshot_at: _dt.datetime,
) -> None:
    """Aba auxiliar com metadados — 1ª aba do workbook (à esquerda)."""
    bold = Font(bold=True)

    ws.cell(row=1, column=1, value="Snapshot do Notion — RPADV").font = bold
    ws.cell(
        row=2, column=1,
        value=(
            "Este arquivo é um instantâneo das bases no momento da geração; "
            "as bases continuam sendo modificadas no Notion após esta exportação."
        ),
    )
    ws.cell(
        row=3, column=1,
        value=f"Gerado em: {snapshot_at.isoformat(timespec='seconds')}",
    )

    ws.cell(row=5, column=1, value="Contagem de registros por base").font = bold
    ws.cell(row=5, column=2, value="Total").font = bold
    row = 6
    for base in bases_in_order:
        ws.cell(row=row, column=1, value=base)
        ws.cell(row=row, column=2, value=counts.get(base, 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Convenções de formatação").font = bold
    row += 1
    legend = [
        "Datas: tipo date nativo do Excel (formato yyyy-mm-dd), exibido na "
        "regional do sistema. Sem formatação cromática.",
        "Checkbox (ex.: Adiantamento): 'Sim' quando marcado no Notion, "
        "vazio caso contrário.",
        "Selects: nome da opção como texto, sem fundo colorido.",
        "Multi-selects: valores separados por vírgula em uma única célula.",
        "Relations: títulos das páginas relacionadas (resolvidos via UUID).",
        "People: nomes do escritório quando reconhecido pelo app; UUIDs do "
        "Notion caso contrário.",
    ]
    for line in legend:
        ws.cell(row=row, column=1, value=line)
        row += 1


def export_snapshot(
    *,
    token: str | None = None,
    client: Any = None,
    bases: list[str],
    dest_path: str,
    on_progress: Callable[[str, str, int, int], None] | None = None,
    schema_registry: Any = None,
    notion_users: dict[str, dict[str, str]] | None = None,
    data_sources: dict[str, str] | None = None,
) -> ExportResult:
    """Round 4 Frente 4 — gera xlsx snapshot das bases selecionadas.

    Args:
        token: token de auth pro NotionClient. Ignorado se ``client``
            for fornecido (útil pra testes).
        client: instância pré-construída de NotionClient (ou mock).
        bases: lista de base_labels a exportar (ex: ["Clientes",
            "Tarefas"]). Bases não conhecidas em ``data_sources`` são
            silenciosamente ignoradas.
        dest_path: caminho absoluto pro arquivo .xlsx de saída.
        on_progress: callback ``(base, phase, count, total)`` opcional.
            ``phase`` é "fetch" ou "write"; ``total`` é -1 quando
            desconhecido (durante fetch).
        schema_registry: instância pra consultar schemas. Default é o
            singleton via ``get_schema_registry()``.
        notion_users: mapping UUID→dict pra resolver people. Default:
            ``notion_bulk_edit.config.NOTION_USERS``.
        data_sources: mapping base→data_source_id. Default:
            ``notion_bulk_edit.config.DATA_SOURCES``.

    Returns:
        ``ExportResult`` com path de destino, contagens por base,
        duração e número de UUIDs de relations não resolvidos (úteis
        pra warning na UI).
    """
    if client is None:
        if not token:
            raise ValueError("export_snapshot precisa de client OU token")
        from notion_bulk_edit.notion_api import NotionClient
        client = NotionClient(token)

    if schema_registry is None:
        from notion_bulk_edit.schema_registry import get_schema_registry
        schema_registry = get_schema_registry()

    if data_sources is None:
        from notion_bulk_edit.config import DATA_SOURCES
        data_sources = DATA_SOURCES

    if notion_users is None:
        from notion_bulk_edit.config import NOTION_USERS
        notion_users = NOTION_USERS

    started = time.time()
    snapshot_at = _dt.datetime.now()

    # Phase 1: fetch all selected bases
    pages_by_base: dict[str, list[dict[str, Any]]] = {}
    bases_in_order: list[str] = []
    for base in bases:
        dsid = data_sources.get(base)
        if not dsid:
            continue
        bases_in_order.append(base)

        def make_progress(b: str):
            def cb(n: int) -> None:
                if on_progress is not None:
                    on_progress(b, PHASE_FETCH, n, -1)
            return cb

        pages = client.query_all(dsid, on_progress=make_progress(base))
        pages_by_base[base] = pages

    # Phase 2: build relation title cache
    title_cache = _build_title_cache(pages_by_base)

    # Phase 3: write workbook
    wb = Workbook()
    # Workbook abre com uma aba default; remova pra ficar só nossas abas.
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    aux_ws = wb.create_sheet(_AUX_SHEET_NAME, 0)

    counts: dict[str, int] = {}
    relation_misses = 0
    for base in bases_in_order:
        pages = pages_by_base.get(base, [])
        ws = wb.create_sheet(title=base)
        miss = _write_base_sheet(
            ws, base, pages, title_cache, notion_users,
            schema_registry, on_progress,
        )
        counts[base] = len(pages)
        relation_misses += miss

    _write_aux_sheet(aux_ws, counts, bases_in_order, snapshot_at)

    wb.save(dest_path)

    return ExportResult(
        dest_path=dest_path,
        counts=counts,
        duration_seconds=time.time() - started,
        relation_misses=relation_misses,
    )
