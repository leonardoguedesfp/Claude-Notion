"""Writer xlsx pras publicações DJEN coletadas pelo ``dje_client``.

Fase 2: pipeline ``dje_transform.transform_rows`` é aplicado ANTES da
escrita — dedup por id, observacoes derivadas das regras A+B, strip
de HTML em ``texto``, normalização de encoding misto em campos
all-caps, drop de 6 colunas redundantes, sort por tribunal+data,
schema canônico de 20 colunas em ordem fixa.

Caller (UI page) continua passando rows brutos do client (formato F1
com ``advogado_consultado`` singular por linha) — toda transformação
é interna ao writer.

Naming: ``Publicacoes_DJEN_{dd.mm.aa_inicio}_a_{dd.mm.aa_fim}_v{N}.xlsx``.
Versão auto-incrementa pra nunca sobrescrever — útil quando o operador
roda a mesma data 2× no mesmo dia (refinou critério).

Fase 2.1 (2026-05-01): defesa em 2 camadas contra ``IllegalCharacterError``
do openpyxl:

1. **Defesa primária** (no ``dje_transform``): sanitiza top-level
   strings antes da escrita (``sanitize_for_xlsx`` no ``_enrich_row``).

2. **Defesa secundária** (aqui):
   - ``_serialize_cell`` aplica ``sanitize_for_xlsx`` na string final
     PÓS-JSON — captura caracteres aninhados em ``destinatarios``
     (lista de dicts) que escaparam da defesa primária.
   - ``write_publicacoes_xlsx`` ganha try/except por linha. Se uma
     célula da linha N falhar, a linha N inteira é pulada (não meia
     linha quebrada) e contabilizada em ``ExportResult.skipped``.
     Demais linhas seguem.

Contrato de retorno mudou em Fase 2.1: ``write_publicacoes_xlsx`` agora
retorna ``ExportResult(path, skipped)`` em vez de ``Path``. Único caller
em produção (``pages/leitor_dje._DJEWorker.run``) consome o NamedTuple
e propaga ``skipped_count`` pra UI exibir banner final composto.

Fase 3 (2026-05-02):
- ``write_publicacoes_xlsx`` ganha kwargs ``oabs_escritorio_marcadas`` e
  ``oabs_externas_pesquisadas`` que são forwarded pro ``transform_rows``
  pra suportar modo manual da UI.
- Schema canônico passa a ter 21 colunas (rename ``advogados_consultados``
  → ``advogados_consultados_escritorio``, nova ``oabs_externas_consultadas``).
- ``write_historico_completo_xlsx`` (novo): reescreve do zero o
  ``Historico_DJEN_completo.xlsx`` (path fixo) a partir de rows do SQLite.
  Estratégia atômica (.tmp + rename) e tolerância a ``PermissionError``
  (arquivo aberto no Excel) — não derruba a execução.
"""
from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path
from typing import Any, NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Font

from notion_rpadv.services.dje_transform import sanitize_for_xlsx

logger = logging.getLogger("dje.exporter")

SHEET_NAME: str = "Publicacoes"
SHEET_STATUS: str = "Status"
SHEET_LOG: str = "Log"
ANNOTATION_COLUMN: str = "advogado_consultado"

# Fase 3 — histórico completo: nome fixo, sobrescreve a cada execução.
HISTORICO_FILENAME: str = "Historico_DJEN_completo.xlsx"
HISTORICO_TMP_FILENAME: str = "Historico_DJEN_completo.tmp.xlsx"

# Schema das colunas das abas auxiliares (refator pós-Fase 3 hotfix).
STATUS_COLUMNS: tuple[str, ...] = (
    "advogado",
    "oab_uf",
    "ultimo_cursor",
    "dias_atras",
    "ultima_execucao",
)
LOG_COLUMNS: tuple[str, ...] = ("mensagem",)


class SkippedRow(NamedTuple):
    """Linha pulada na escrita do xlsx (Fase 2.1)."""

    source_idx: int   # 1-based index na lista processed_rows
    row_id: Any       # ``row['id']`` se existir, senão None
    error: str        # ``str(exc)`` truncado pra log


class ExportResult(NamedTuple):
    """Resultado de ``write_publicacoes_xlsx`` (Fase 2.1).

    ``path``: caminho final do arquivo gerado.
    ``skipped``: linhas que falharam ao escrever — vazio em sucesso total.
    Caller pode iterar pra exibir detalhes ou usar ``len(skipped)`` pra
    banner final.
    """

    path: Path
    skipped: list[SkippedRow]


def format_filename(
    data_inicio: date, data_fim: date, version: int,
) -> str:
    """``Publicacoes_DJEN_{dd.mm.aa}_a_{dd.mm.aa}_v{N}.xlsx``."""
    di = data_inicio.strftime("%d.%m.%y")
    df = data_fim.strftime("%d.%m.%y")
    return f"Publicacoes_DJEN_{di}_a_{df}_v{version}.xlsx"


def next_version(
    output_dir: Path, data_inicio: date, data_fim: date,
) -> int:
    """Próxima versão livre pra esse intervalo. v1 quando nada existe;
    v2/v3/... quando arquivos anteriores presentes."""
    v = 1
    while (output_dir / format_filename(data_inicio, data_fim, v)).exists():
        v += 1
    return v


def _resolve_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Ordem de colunas: ``advogado_consultado`` primeiro, depois as
    chaves do JSON da API na ordem em que aparecem no PRIMEIRO item
    retornado (defesa contra schemas que misturam ordem entre chamadas).

    Itens posteriores podem ter chaves novas (defesa contra mudanças do
    DJEN) — essas vão pro fim da lista, na ordem de aparição.
    """
    if not rows:
        return [ANNOTATION_COLUMN]
    seen: set[str] = {ANNOTATION_COLUMN}
    columns: list[str] = [ANNOTATION_COLUMN]
    # Primeiro item dita a ordem principal das chaves do JSON.
    for k in rows[0].keys():
        if k not in seen:
            columns.append(k)
            seen.add(k)
    # Itens subsequentes contribuem só com chaves novas, no final.
    for row in rows[1:]:
        for k in row.keys():
            if k not in seen:
                columns.append(k)
                seen.add(k)
    return columns


def _serialize_cell(value: Any) -> Any:
    """Converte valor decoded pra algo que openpyxl aceita.

    - Listas/dicts → JSON string com ensure_ascii=False (preserva acentos)
      → sanitiza string final via ``sanitize_for_xlsx`` (Fase 2.1: defesa
      secundária pra control chars aninhados em ``destinatarios``).
    - Strings → sanitiza (idempotente — defesa em depth se transform
      foi pulado por algum motivo).
    - None → None (célula vazia).
    - Demais → passa direto.
    """
    if value is None:
        return None
    if isinstance(value, (list, dict)):
        return sanitize_for_xlsx(json.dumps(value, ensure_ascii=False))
    if isinstance(value, str):
        return sanitize_for_xlsx(value)
    return value


class HistoricoResult(NamedTuple):
    """Resultado de ``write_historico_completo_xlsx`` (Fase 3).

    ``path``: caminho final do .xlsx, ou ``None`` se a escrita foi
    pulada por arquivo bloqueado.
    ``skipped``: linhas que falharam ao escrever (mesma defesa F2.1).
    ``locked``: ``True`` se o arquivo destino estava bloqueado pra
    escrita (Windows file lock — usuário com Excel aberto). Nesse caso,
    ``path=None`` e o arquivo anterior permanece intacto. Caller
    (worker) registra warning e segue sem derrubar a execução.
    """

    path: Path | None
    skipped: list[SkippedRow]
    locked: bool


def _format_date_br(d) -> str:
    """``2026-04-30`` → ``"30/04/2026"`` ou ``"—"`` se ``None``."""
    if d is None:
        return "—"
    return d.strftime("%d/%m/%Y")


def _format_datetime_br(dt) -> str:
    if dt is None:
        return "—"
    return dt.strftime("%d/%m/%Y %H:%M")


def _add_aba_status(
    wb,
    advogados: list[dict],
    state_map: dict,
    *,
    today=None,
) -> None:
    """Adiciona a aba "Status" com 1 linha por advogado oficial.

    ``state_map`` vem de ``dje_state.read_all_advogados_state(conn)`` —
    dict ``(oab, uf) → {ultimo_cursor, last_run}``. Advogados sem state
    aparecem com cursor "—".
    """
    from datetime import date as _date

    if today is None:
        today = _date.today()
    ws = wb.create_sheet(SHEET_STATUS)
    bold = Font(bold=True)
    for col_idx, name in enumerate(STATUS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
    for row_idx, adv in enumerate(advogados, start=2):
        st = state_map.get((adv["oab"], adv["uf"]), {})
        cursor = st.get("ultimo_cursor")
        last_run = st.get("last_run")
        dias_atras = (today - cursor).days if cursor is not None else "—"
        ws.cell(row=row_idx, column=1, value=adv["nome"])
        ws.cell(row=row_idx, column=2, value=f"{adv['oab']}/{adv['uf']}")
        ws.cell(row=row_idx, column=3, value=_format_date_br(cursor))
        ws.cell(row=row_idx, column=4, value=dias_atras)
        ws.cell(row=row_idx, column=5, value=_format_datetime_br(last_run))


def _add_aba_log(wb, log_lines: list[str]) -> None:
    """Adiciona a aba "Log" com 1 linha por mensagem do log da execução.

    Mesma informação que aparece no painel de log da UI durante a
    varredura, persistida pra auditoria offline.
    """
    ws = wb.create_sheet(SHEET_LOG)
    bold = Font(bold=True)
    for col_idx, name in enumerate(LOG_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
    for row_idx, line in enumerate(log_lines, start=2):
        ws.cell(row=row_idx, column=1, value=sanitize_for_xlsx(line))


def _write_workbook(
    processed_rows: list[dict[str, Any]],
    columns: list[str],
    target_path: Path,
    *,
    advogados: list[dict] | None = None,
    state_map: dict | None = None,
    log_lines: list[str] | None = None,
) -> list[SkippedRow]:
    """Helper interno: escreve um workbook xlsx com a defesa per-row
    da Fase 2.1 e salva em ``target_path``.

    Retorna lista de ``SkippedRow`` (vazia em sucesso total).

    Levanta ``PermissionError`` se ``target_path`` está bloqueado pra
    escrita (caller decide se trata).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    bold = Font(bold=True)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold

    skipped: list[SkippedRow] = []
    written_row_idx = 1  # row 1 = header
    for source_idx, row in enumerate(processed_rows, start=1):
        # Próximo slot disponível. Se a linha falhar, NÃO avança —
        # a próxima linha (sucesso) sobrescreve as células zeradas
        # do rollback, mantendo o arquivo sem gaps.
        target_row_idx = written_row_idx + 1
        cells_set: list[int] = []
        write_error: Exception | None = None
        for col_idx, key in enumerate(columns, start=1):
            value = _serialize_cell(row.get(key))
            if value is None:
                continue
            try:
                ws.cell(row=target_row_idx, column=col_idx, value=value)
                cells_set.append(col_idx)
            except Exception as exc:  # noqa: BLE001
                write_error = exc
                break
        if write_error is not None:
            # Rollback: zera as células parcialmente escritas.
            for c in cells_set:
                ws.cell(row=target_row_idx, column=c).value = None
            sk = SkippedRow(
                source_idx=source_idx,
                row_id=row.get("id"),
                error=str(write_error)[:200],
            )
            skipped.append(sk)
            logger.warning(
                "DJE: linha pulada na escrita: source_idx=%d id=%r erro=%s",
                sk.source_idx, sk.row_id, sk.error,
            )
        else:
            written_row_idx = target_row_idx

    # Abas auxiliares (refator pós-Fase 3 hotfix): "Status" com 1 linha
    # por advogado oficial + "Log" da execução. Caller passa ``advogados``
    # e ``state_map``/``log_lines`` quando aplicável; sem esses params,
    # as abas não são criadas (compat com legado).
    if advogados is not None and state_map is not None:
        _add_aba_status(wb, advogados, state_map)
    if log_lines is not None:
        _add_aba_log(wb, log_lines)

    wb.save(target_path)
    return skipped


def write_publicacoes_xlsx(
    rows: list[dict[str, Any]],
    output_dir: Path,
    data_inicio: date,
    data_fim: date,
    *,
    oabs_escritorio_marcadas: set[str] | None = None,
    oabs_externas_pesquisadas: set[str] | None = None,
) -> ExportResult:
    """Escreve o xlsx no diretório informado, com versionamento
    automático. Retorna ``ExportResult(path, skipped)``.

    ``output_dir`` é criado se não existir (mas se o caller esperou
    fallback de QFileDialog quando ``output_dir`` é inacessível,
    deve ter resolvido isso ANTES de chamar — esta função só faz
    mkdir simples, sem prompt).

    Fase 2: aplica ``dje_transform.transform_rows`` antes de escrever
    — dedup por id, observacoes A+B, strip HTML, normalização de
    encoding, drop de colunas redundantes, sort, schema canônico de
    21 colunas (Fase 3). Caller continua passando rows brutos do client.

    Fase 2.1: try/except por linha. Célula que falhar (e.g.
    ``IllegalCharacterError`` que escapou da defesa primária) faz a
    linha INTEIRA ser pulada — não meia linha quebrada — e
    contabilizada em ``result.skipped``. Demais linhas seguem normalmente.

    Fase 3: kwargs ``oabs_escritorio_marcadas`` e ``oabs_externas_pesquisadas``
    são forwarded pro ``transform_rows`` pra suportar o split do modo
    manual. Defaults preservam o comportamento do modo padrão.
    """
    from notion_rpadv.services.dje_transform import transform_rows
    processed_rows, columns = transform_rows(
        rows,
        oabs_escritorio_marcadas=oabs_escritorio_marcadas,
        oabs_externas_pesquisadas=oabs_externas_pesquisadas,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    version = next_version(output_dir, data_inicio, data_fim)
    path = output_dir / format_filename(data_inicio, data_fim, version)

    skipped = _write_workbook(processed_rows, columns, path)
    logger.info(
        "DJE: xlsx salvo em %s (%d linhas escritas, %d colunas) — "
        "raw=%d → deduped=%d, %d puladas",
        path, len(processed_rows) - len(skipped), len(columns),
        len(rows), len(processed_rows), len(skipped),
    )
    return ExportResult(path=path, skipped=skipped)


def write_publicacoes_xlsx_from_processed(
    processed_rows: list[dict[str, Any]],
    output_dir: Path,
    data_inicio: date,
    data_fim: date,
    *,
    advogados: list[dict] | None = None,
    state_map: dict | None = None,
    log_lines: list[str] | None = None,
) -> ExportResult:
    """Escreve Excel-de-execução versionado a partir de rows JÁ
    pós-split (com ``advogados_consultados_escritorio`` e
    ``oabs_externas_consultadas`` calculados).

    Diferença vs ``write_publicacoes_xlsx``: pula dedup + split, faz só
    enrich (observacoes, sanitize, etc) + sort. Usado pelo
    ``_DJEWorker`` no fluxo Fase 3 que faz dedup/split inline pra
    armazenar no SQLite e reaproveita as rows pra gerar o Excel.

    Refator pós-Fase 3 hotfix: kwargs ``advogados``, ``state_map`` e
    ``log_lines`` opcionais — quando passados, gera abas auxiliares
    "Status" (1 linha por advogado oficial com cursor + dias atrás)
    e "Log" (mensagens da execução).

    Refator pós-Fase 3: ``processed_rows`` pode ser lista vazia —
    nesse caso o Excel é gerado mesmo assim (só cabeçalho + abas
    Status/Log) pra que o usuário sempre tenha evidência da execução.

    Mesma estratégia de versionamento e mesma defesa per-row do F2.1.
    """
    from notion_rpadv.services.dje_transform import (
        transform_rows_for_history,
    )

    enriched_rows, columns = transform_rows_for_history(processed_rows)

    output_dir.mkdir(parents=True, exist_ok=True)
    version = next_version(output_dir, data_inicio, data_fim)
    path = output_dir / format_filename(data_inicio, data_fim, version)

    skipped = _write_workbook(
        enriched_rows, columns, path,
        advogados=advogados, state_map=state_map, log_lines=log_lines,
    )
    logger.info(
        "DJE: xlsx (de execução, pós-split) salvo em %s "
        "(%d linhas escritas, %d colunas, %d puladas)",
        path, len(enriched_rows) - len(skipped), len(columns), len(skipped),
    )
    return ExportResult(path=path, skipped=skipped)


def write_historico_completo_xlsx(
    db_rows: list[dict[str, Any]],
    output_dir: Path,
    *,
    advogados: list[dict] | None = None,
    state_map: dict | None = None,
    log_lines: list[str] | None = None,
) -> HistoricoResult:
    """Reescreve o ``Historico_DJEN_completo.xlsx`` (path fixo) a partir
    das rows do SQLite.

    ``db_rows`` deve vir de ``dje_db.fetch_all_publicacoes()`` — cada row
    já contém ``advogados_consultados_escritorio`` e
    ``oabs_externas_consultadas`` pré-calculados, então pula dedup+split
    e faz só enrich (observacoes, sanitize, etc) + sort
    (``transform_rows_for_history``).

    Estratégia de escrita atômica: escreve em ``.tmp.xlsx``, depois faz
    rename pra substituir o final. Se o ``.tmp`` falhar com
    ``PermissionError`` (Windows: arquivo aberto no Excel), retorna
    ``HistoricoResult(path=None, locked=True)`` sem derrubar a execução.
    Arquivo anterior permanece intacto.

    Caller (``_DJEWorker.run``) loga warning quando ``locked=True``.
    """
    from notion_rpadv.services.dje_transform import (
        transform_rows_for_history,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    final_path = output_dir / HISTORICO_FILENAME
    tmp_path = output_dir / HISTORICO_TMP_FILENAME

    processed_rows, columns = transform_rows_for_history(db_rows)

    try:
        skipped = _write_workbook(
            processed_rows, columns, tmp_path,
            advogados=advogados, state_map=state_map, log_lines=log_lines,
        )
    except PermissionError as exc:
        logger.warning(
            "DJE: %s bloqueado pra escrita (provável arquivo aberto no Excel): %s",
            tmp_path, exc,
        )
        return HistoricoResult(path=None, skipped=[], locked=True)

    # Rename atômico do .tmp pro path final. Em Windows, replace = atômico
    # quando ambos estão na mesma volume e o destino existe.
    try:
        tmp_path.replace(final_path)
    except PermissionError as exc:
        logger.warning(
            "DJE: %s bloqueado pra rename (provável arquivo aberto no Excel): "
            "%s — .tmp em %s permanece como evidência",
            final_path, exc, tmp_path,
        )
        return HistoricoResult(path=None, skipped=skipped, locked=True)

    logger.info(
        "DJE: histórico salvo em %s (%d linhas, %d puladas)",
        final_path, len(processed_rows) - len(skipped), len(skipped),
    )
    return HistoricoResult(path=final_path, skipped=skipped, locked=False)
